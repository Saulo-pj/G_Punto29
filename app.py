import os
import importlib
import json
import re
import secrets
import uuid
import hmac
from io import BytesIO
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from collections import defaultdict
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, url_for, session
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from dotenv import load_dotenv
from sqlalchemy import or_
from sqlalchemy import inspect, text
from sqlalchemy.exc import IntegrityError
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from database import (
	db,
	Sede,
	Rol,
	Turno,
	RecordatorioCierre,
	Categoria,
	Unidad,
	Area,
	Subarea,
	Usuario,
	Producto,
	InventarioSede,
	ChecklistPedido,
	DetallePedido,
	PlantillaChecklistItem,
	MovimientoInventario,
	ArqueoCaja,
	ArqueoCajaHistorial,
	Merma,
	Incidencia,
	CatalogoMerma,
	AgendaAuditoria,
	AgendaPersistencia,
)

# Inicializar LoginManager global para decoradores y uso antes de create_app
login_manager = LoginManager()
login_manager.login_view = 'login'

DEFAULT_AREAS = {
	'cocina': ['cocina_caliente', 'cocina_fria'],
	'sala': ['sala'],
}


def _get_operation_date():
    # Fecha/hora base de operacion; simplificada a ahora local
    return datetime.now()


def _get_selected_app_date():
    selected_date = session.get('app_date', '').strip()
    if not selected_date:
        operation_date = _get_operation_date().date()
        session['app_date'] = operation_date.strftime('%Y-%m-%d')
        return operation_date
    try:
        return datetime.strptime(selected_date, '%Y-%m-%d').date()
    except ValueError:
        operation_date = _get_operation_date().date()
        session['app_date'] = operation_date.strftime('%Y-%m-%d')
        return operation_date


def _format_peru_datetime(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except ValueError:
            return str(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    peru_tz = ZoneInfo('America/Lima')
    return dt.astimezone(peru_tz).strftime('%d/%m/%Y %H:%M')


def _allowed_views(user):
	view_order = [
		('dashboard', 'Inicio'),
		('inventario', 'Inventario'),
		('movimientos', 'Movimientos'),
		('pedidos', 'Pedidos'),
		('checklist', 'Checklist'),
		('arqueo', 'Arqueo Caja'),
		('mermas', 'Mermas'),
		('incidencias', 'Incidencias'),
		('horarios', 'Gestion de Personal y Horarios'),
		('ajustes', 'Ajustes'),
	]
	allowed = [item for item in view_order if user.can_view(item[0])]
	if user.can_view('inventario'):
		insert_index = 1
		for idx, item in enumerate(allowed):
			if item[0] == 'inventario':
				insert_index = idx + 1
				break
		allowed.insert(insert_index, ('inventario_dashboard', 'Dashboard Inventario'))
	return allowed


def _seed_catalogs():
	if Rol.query.count() == 0:
		for name in ['admin_general', 'admin_almacen', 'personal_prod', 'admin_sala', 'cocinero']:
			db.session.add(Rol(nombre_rol=name))

	for default_sede in ('Almacen', 'Sede_17', 'Sede_20'):
		if not Sede.query.filter(db.func.lower(Sede.nombre_sede) == default_sede.lower()).first():
			db.session.add(Sede(nombre_sede=default_sede))

	if Turno.query.count() == 0:
		for code, name in [('MANANA', 'Manana'), ('NOCHE', 'Noche'), ('NA', 'N/A')]:
			db.session.add(Turno(id_turno=code, nombre_turno=name))

	if Categoria.query.count() == 0:
		for name in ['Carnes', 'Pollos', 'Condimentos', 'Abarrotes', 'Preparados']:
			db.session.add(Categoria(nombre_categoria=name))
	else:
		legacy_categories = ['Cocina', 'Sala', 'Almacen']
		for legacy in legacy_categories:
			if not Producto.query.filter(Producto.id_area == legacy).first():
				legacy_item = Categoria.query.filter_by(nombre_categoria=legacy).first()
				if legacy_item:
					db.session.delete(legacy_item)
		for name in ['Carnes', 'Pollos', 'Condimentos', 'Abarrotes', 'Preparados']:
			if not Categoria.query.filter(db.func.lower(Categoria.nombre_categoria) == name.lower()).first():
				db.session.add(Categoria(nombre_categoria=name))

	if Unidad.query.count() == 0:
		for name in ['kg', 'Litro', 'unidad']:
			db.session.add(Unidad(nombre_unidad=name))

	if Area.query.count() == 0:
		for name in ['cocina', 'sala']:
			db.session.add(Area(nombre_area=name))

	db.session.flush()
	for area_name, subareas in DEFAULT_AREAS.items():
		area = Area.query.filter(db.func.lower(Area.nombre_area) == area_name).first()
		if area:
			for subarea_name in subareas:
				if not Subarea.query.filter_by(id_area=area.id_area, nombre_subarea=subarea_name).first():
					db.session.add(Subarea(id_area=area.id_area, nombre_subarea=subarea_name))

	db.session.commit()

	admin_role = Rol.query.filter_by(nombre_rol='admin_general').first()
	main_sede = Sede.query.first()
	na_turno = Turno.query.filter_by(id_turno='NA').first()
	admin_user = Usuario.query.filter_by(username='admin').first()
	if not admin_user and admin_role and main_sede and na_turno:
		db.session.add(
			Usuario(
				id_usuario='admin',
				username='admin',
				password_hash=generate_password_hash('admin1234'),
				id_rol=admin_role.id_rol,
				id_sede=main_sede.id_sede,
				id_turno=na_turno.id_turno,
			)
		)
		db.session.commit()


@login_manager.user_loader
def load_user(user_id):
	return db.session.get(Usuario, user_id)


def _forbidden_redirect():
	flash('No tienes permisos para entrar a esta vista.', 'error')
	return redirect(url_for('dashboard'))


def _safe_json_list(value):
	if value in (None, '', 'null'):
		return []
	try:
		payload = json.loads(value)
		return payload if isinstance(payload, list) else []
	except (TypeError, ValueError):
		return []


def _stats_for_user(user):
	try:
		stats = {
			'productos': Producto.query.count(),
			'movimientos': MovimientoInventario.query.count(),
			'pedidos': ChecklistPedido.query.count(),
			'arqueos': ArqueoCaja.query.count(),
		}
		return stats
	except Exception:
		return {'productos': 0, 'movimientos': 0, 'pedidos': 0, 'arqueos': 0}


def _home_alerts_for_user(user, selected_date):
	try:
		date_str = selected_date.strftime('%Y-%m-%d') if selected_date else datetime.utcnow().strftime('%Y-%m-%d')
		alerts = {
			'stock_critico_count': 0,
			'pedidos_pendientes_count': 0,
			'missing_arqueo': False,
			'subtitle': 'Resumen de tareas para hoy segun tu rol.',
			'cards': [],
		}

		stock_query = InventarioSede.query
		if user and user.rol_nombre != 'admin_general':
			stock_query = stock_query.filter(InventarioSede.id_sede == user.id_sede)
		alerts['stock_critico_count'] = stock_query.filter(
			InventarioSede.punto_minimo > 0,
			InventarioSede.stock_actual <= InventarioSede.punto_minimo,
		).count()

		if user and user.can_view('pedidos'):
			pedidos_query = ChecklistPedido.query.filter(
				db.func.date(ChecklistPedido.fecha) == date_str,
				ChecklistPedido.estado_general == 'Pendiente',
			)
			if user.rol_nombre not in {'admin_general', 'admin_almacen', 'personal_prod'}:
				pedidos_query = pedidos_query.filter(ChecklistPedido.id_sede == user.id_sede)
			alerts['pedidos_pendientes_count'] = pedidos_query.count()

		if user and user.can_view('arqueo'):
			arqueo_query = ArqueoCaja.query.filter(ArqueoCaja.fecha == selected_date)
			if user.rol_nombre != 'admin_general':
				arqueo_query = arqueo_query.filter(
					ArqueoCaja.id_sede == user.id_sede,
					ArqueoCaja.id_turno == user.id_turno,
				)
			alerts['missing_arqueo'] = arqueo_query.count() == 0

		role_name = user.rol_nombre if user else ''

		if role_name == 'admin_general':
			alerts['subtitle'] = 'Vision global: pendientes de todo el equipo.'
			checklists_pendientes = ChecklistPedido.query.filter(
				db.func.date(ChecklistPedido.fecha) == date_str,
				ChecklistPedido.estado_general.in_(['Borrador', 'Pendiente']),
			).count()
			admin_sala_scopes = db.session.query(Usuario.id_sede, Usuario.id_turno).join(
				Rol, Rol.id_rol == Usuario.id_rol
			).filter(
				Rol.nombre_rol == 'admin_sala'
			).distinct().all()
			missing_arqueos_count = 0
			for scope in admin_sala_scopes:
				if ArqueoCaja.query.filter_by(
					fecha=selected_date,
					id_sede=scope.id_sede,
					id_turno=scope.id_turno,
				).first() is None:
					missing_arqueos_count += 1

			alerts['cards'] = [
				{'title': 'Stock critico', 'message': f"Hay {alerts['stock_critico_count']} productos con stock critico.", 'state': 'warn' if alerts['stock_critico_count'] > 0 else 'ok', 'link': 'inventario'},
				{'title': 'Pedidos pendientes', 'message': f"Hay {alerts['pedidos_pendientes_count']} pedidos sin cerrar.", 'state': 'warn' if alerts['pedidos_pendientes_count'] > 0 else 'ok', 'link': 'pedidos'},
				{'title': 'Checklist pendientes', 'message': f"Hay {checklists_pendientes} checklist en borrador o pendiente.", 'state': 'warn' if checklists_pendientes > 0 else 'ok', 'link': 'checklist'},
				{'title': 'Arqueos faltantes', 'message': f"Faltan {missing_arqueos_count} arqueos de admin sala por registrar.", 'state': 'warn' if missing_arqueos_count > 0 else 'ok', 'link': 'arqueo'},
			]
			return alerts

		if role_name == 'cocinero':
			alerts['subtitle'] = 'Tu foco es completar y enviar tu checklist del turno.'
			my_checklist = _checklist_base_query(user, selected_date).order_by(ChecklistPedido.id_pedido.desc()).first()
			checklist_done = bool(my_checklist and my_checklist.estado_general in {'Enviado', 'Finalizado'})
			if my_checklist is None:
				message = 'Aun no creaste tu lista de hoy.'
			elif checklist_done:
				message = f'Tu lista ya fue enviada. Estado: {my_checklist.estado_general}.'
			else:
				message = f"Tu lista aun esta en estado {my_checklist.estado_general}."
			alerts['cards'] = [{'title': 'Checklist de cocina', 'message': message, 'state': 'ok' if checklist_done else 'warn', 'link': 'checklist'}]
			return alerts

		if role_name == 'admin_sala':
			alerts['subtitle'] = 'Hoy debes completar checklist y registrar arqueo de caja.'
			my_checklist = _checklist_base_query(user, selected_date).order_by(ChecklistPedido.id_pedido.desc()).first()
			checklist_done = bool(my_checklist and my_checklist.estado_general in {'Enviado', 'Finalizado'})
			if my_checklist is None:
				checklist_message = 'Aun no creaste tu checklist del turno.'
			elif checklist_done:
				checklist_message = 'Checklist completado correctamente.'
			else:
				checklist_message = f"Checklist en progreso ({my_checklist.estado_general})."
			arqueo_done = not alerts['missing_arqueo']
			alerts['cards'] = [
				{'title': 'Checklist de sala', 'message': checklist_message, 'state': 'ok' if checklist_done else 'warn', 'link': 'checklist'},
				{'title': 'Arqueo de caja', 'message': 'Arqueo registrado para hoy.' if arqueo_done else 'Falta registrar arqueo de caja hoy.', 'state': 'ok' if arqueo_done else 'warn', 'link': 'arqueo'},
			]
			return alerts

		if role_name == 'personal_prod':
			alerts['subtitle'] = 'Seguimiento de tu lista de pedidos de produccion.'
			my_pedido = ChecklistPedido.query.filter(
				ChecklistPedido.id_usuario == user.id_usuario,
				db.func.date(ChecklistPedido.fecha) == date_str,
			).order_by(ChecklistPedido.id_pedido.desc()).first()
			created = my_pedido is not None
			sent = bool(my_pedido and my_pedido.estado_general in {'Enviado', 'Finalizado'})
			alerts['cards'] = [
				{'title': 'Lista creada', 'message': 'Tu lista de pedidos de hoy ya existe.' if created else 'Todavia no creaste tu lista de pedidos de hoy.', 'state': 'ok' if created else 'warn', 'link': 'pedidos'},
				{'title': 'Lista enviada', 'message': 'Tu lista ya fue enviada a sede.' if sent else 'Aun no enviaste la lista a sede.', 'state': 'ok' if sent else 'warn', 'link': 'pedidos'},
			]
			return alerts

		alerts['cards'] = [
			{'title': 'Stock bajo', 'message': f"Hay {alerts['stock_critico_count']} productos con stock critico.", 'state': 'warn' if alerts['stock_critico_count'] > 0 else 'ok', 'link': 'inventario'},
			{'title': 'Pedidos pendientes', 'message': f"Tienes {alerts['pedidos_pendientes_count']} pedidos por revisar.", 'state': 'warn' if alerts['pedidos_pendientes_count'] > 0 else 'ok', 'link': 'pedidos'},
		]
		return alerts
	except Exception:
		return {
			'stock_critico_count': 0,
			'pedidos_pendientes_count': 0,
			'missing_arqueo': False,
			'subtitle': 'Resumen de tareas para hoy segun tu rol.',
			'cards': [{'title': 'Sin pendientes', 'message': 'No se pudo calcular el resumen del dashboard. Reintenta.', 'state': 'ok', 'link': None}],
		}


def _inventory_dashboard_metrics(user, selected_date):
	base_inv = db.session.query(Producto, InventarioSede, Sede).join(
		InventarioSede,
		Producto.id_producto == InventarioSede.id_producto,
	).outerjoin(
		Sede,
		Sede.id_sede == InventarioSede.id_sede,
	)
	if user.rol_nombre != 'admin_general':
		base_inv = base_inv.filter(InventarioSede.id_sede == user.id_sede)

	inventory_rows = base_inv.order_by(Producto.nombre_producto.asc()).all()
	por_acabarse = []
	acabados = []
	abastecidos = 0

	for producto, inv, sede in inventory_rows:
		stock = _safe_float(inv.stock_actual, 0.0)
		minimo = _safe_float(inv.punto_minimo, 0.0)
		if stock <= 0:
			acabados.append((producto, inv, sede))
		elif minimo > 0 and stock <= minimo:
			por_acabarse.append((producto, inv, sede))
		else:
			abastecidos += 1

	period_start = selected_date - timedelta(days=29)
	salidas_query = db.session.query(
		MovimientoInventario.id_producto,
		Producto.nombre_producto,
		db.func.sum(MovimientoInventario.cantidad).label('total_salida'),
	).join(
		Producto,
		Producto.id_producto == MovimientoInventario.id_producto,
	).filter(
		db.func.upper(MovimientoInventario.tipo) == 'SALIDA',
		db.func.date(MovimientoInventario.fecha) >= period_start.strftime('%Y-%m-%d'),
		db.func.date(MovimientoInventario.fecha) <= selected_date.strftime('%Y-%m-%d'),
	)
	if user.rol_nombre != 'admin_general':
		salidas_query = salidas_query.filter(MovimientoInventario.id_sede == user.id_sede)

	top_salidas = salidas_query.group_by(
		MovimientoInventario.id_producto,
		Producto.nombre_producto,
	).order_by(text('total_salida DESC')).limit(8).all()

	return {
		'kpis': {
			'total_items': len(inventory_rows),
			'por_acabarse': len(por_acabarse),
			'acabados': len(acabados),
			'abastecidos': abastecidos,
		},
		'top_salidas': top_salidas,
		'por_acabarse': por_acabarse[:30],
		'acabados': acabados[:30],
		'chart_top_salidas': {
			'labels': [row.nombre_producto or row.id_producto for row in top_salidas],
			'values': [round(_safe_float(row.total_salida, 0.0), 2) for row in top_salidas],
		},
		'chart_stock': {
			'labels': ['Abastecidos', 'Por acabarse', 'Acabados'],
			'values': [abastecidos, len(por_acabarse), len(acabados)],
		},
	}


def _inventory_query_for_user(user, q='', categoria='', subarea='', unidad='', area=''):
	query = db.session.query(Producto, InventarioSede, Sede).join(
		InventarioSede,
		Producto.id_producto == InventarioSede.id_producto,
	).join(Sede, Sede.id_sede == InventarioSede.id_sede)

	if q:
		like_q = f"%{q}%"
		query = query.filter(
			or_(
				Producto.id_producto.ilike(like_q),
				Producto.nombre_producto.ilike(like_q),
				Producto.id_area.ilike(like_q),
				Producto.area.ilike(like_q),
				Producto.subarea.ilike(like_q),
			)
		)

	if categoria:
		query = query.filter(Producto.id_area == categoria)

	if area:
		query = query.filter(Producto.area == _normalize_area(area))

	if subarea:
		query = query.filter(Producto.subarea == subarea)

	if unidad:
		query = query.filter(Producto.unidad == unidad)

	return query.order_by(Producto.nombre_producto.asc())


def _safe_float(value, default=0.0):
	try:
		return float(value)
	except (TypeError, ValueError):
		return default


def _apply_dispatch_inventory_delta(id_sede, id_producto, delta, pedido_id, detalle_id):
	if not id_sede or not id_producto or delta == 0:
		return

	row = InventarioSede.query.filter_by(id_sede=id_sede, id_producto=id_producto).first()
	if not row:
		row = InventarioSede(id_sede=id_sede, id_producto=id_producto, stock_actual=0.0, punto_minimo=0.0)
		db.session.add(row)

	stock_actual = _safe_float(row.stock_actual, 0.0)
	stock_change = -delta
	new_stock = stock_actual + stock_change
	if new_stock < 0:
		raise ValueError(f'Stock insuficiente para actualizar pedido. Disponible: {stock_actual:.2f}')

	row.stock_actual = new_stock
	db.session.add(
		MovimientoInventario(
			id_sede=id_sede,
			id_producto=id_producto,
			cantidad=abs(delta),
			tipo='SALIDA' if delta > 0 else 'ENTRADA',
			motivo=f'Pedido #{pedido_id} linea #{detalle_id}',
			fecha=datetime.utcnow(),
			id_usuario=current_user.id_usuario,
		)
	)


def _generate_product_id():
	existing_ids = [row[0] for row in db.session.query(Producto.id_producto).all() if row[0]]
	max_number = 0
	for existing_id in existing_ids:
		match = re.match(r'^PROD(\d+)$', existing_id.strip(), re.IGNORECASE)
		if match:
			max_number = max(max_number, int(match.group(1)))

	next_number = max_number + 1
	while True:
		candidate = f'PROD{next_number:04d}'
		if candidate not in existing_ids:
			return candidate
		next_number += 1


def _parse_gastos_from_form(form_data):
	nombres = form_data.getlist('gasto_nombre[]')
	montos = form_data.getlist('gasto_monto[]')
	tipos = form_data.getlist('gasto_tipo[]')
	gastos = []
	allowed_types = {
		'Productos de cocina faltantes', 'Productos de sala faltantes', 'Comida',
		'Propina', 'Marketing', 'Almacén', 'Otros',
	}
	for index, (nombre, monto_raw) in enumerate(zip(nombres, montos)):
		nombre_limpio = (nombre or '').strip()
		monto = _safe_float(monto_raw, 0.0)
		tipo = (tipos[index] if index < len(tipos) else 'Otros').strip()
		if tipo not in allowed_types:
			tipo = 'Otros'
		if not nombre_limpio and monto <= 0:
			continue
		if monto < 0:
			monto = 0.0
		gastos.append({
			'id': f'gasto-{index + 1}',
			'nombre': nombre_limpio or tipo,
			'tipo': tipo,
			'monto': monto,
			'bloqueado': True,
		})
	return gastos


def _normalizar_gastos_arqueo(gastos):
	"""Asegura que cada gasto tenga un identificador persistente y único."""
	normalizados = []
	ids_vistos = set()
	cambio = False
	for gasto in gastos or []:
		if not isinstance(gasto, dict):
			cambio = True
			continue
		item = dict(gasto)
		item_id = str(item.get('id') or '').strip()
		if not item_id or item_id in ids_vistos:
			item_id = f'gasto-{uuid.uuid4().hex}'
			item['id'] = item_id
			cambio = True
		ids_vistos.add(item_id)
		if 'bloqueado' not in item:
			item['bloqueado'] = True
			cambio = True
		normalizados.append(item)
	return normalizados, cambio


def _cambios_gastos_arqueo(anterior, nuevo):
	"""Devuelve sólo las filas de gastos que realmente cambiaron.

	El historial conserva dos columnas (antes/después), por eso una alta se
	registra como ``[] -> [gasto]``, una edición como ``[antes] -> [después]`` y
	una eliminación como ``[gasto] -> []``.
	"""
	antes_por_id = {str(item.get('id')): item for item in anterior if item.get('id')}
	nuevos_por_id = {str(item.get('id')): item for item in nuevo if item.get('id')}
	antes = []
	despues = []
	for item_id in sorted(set(antes_por_id) | set(nuevos_por_id)):
		valor_anterior = antes_por_id.get(item_id)
		valor_nuevo = nuevos_por_id.get(item_id)
		if valor_anterior == valor_nuevo:
			continue
		if valor_anterior is not None:
			antes.append(valor_anterior)
		if valor_nuevo is not None:
			despues.append(valor_nuevo)
	return antes, despues


def _extract_fields_from_form(form_data):
	fields = {}
	for k in ('monto_inicial', 'pos_tarjetas', 'yape', 'plin', 'efectivo', 'venta_sistema', 'observaciones', 'efectivo_entregado', 'efectivo_dejado_caja_real'):
		if k in form_data and form_data.get(k, '').strip() != '':
			fields[k] = form_data.get(k).strip()
	gastos = _parse_gastos_from_form(form_data)
	if gastos:
		fields['gastos'] = gastos
	return fields


def _process_arqueo_save(cierre, payload, is_admin_general, current_user, target_sede_id):
	fields = payload.get('fields') or {}
	is_autosave = payload.get('event') == 'AUTOGUARDADO_5MIN'
	actor_id = getattr(current_user, 'id_usuario', None) or getattr(current_user, 'get_id', lambda: None)()

	def _audit(campo, anterior, nuevo):
		_audit_arqueo_event_with_user(cierre, 'AUTOGUARDADO_5MIN' if is_autosave else 'GUARDADO_MANUAL', campo, anterior, nuevo, actor_id)

	try:
		locked_fields = set(json.loads(cierre.campos_bloqueados_json or '[]'))
	except (TypeError, ValueError):
		locked_fields = set()

	allowed_numeric = {'monto_inicial', 'pos_tarjetas', 'yape', 'plin', 'efectivo', 'venta_sistema'}
	for field in allowed_numeric:
		if field not in fields:
			continue
		raw_value = fields.get(field)
		# Ignorar campos vacíos ('', None) para no bloquearlos ni alterarlos erróneamente
		if raw_value in ('', None):
			continue
		if field in locked_fields and not is_admin_general:
			continue
		old_value = getattr(cierre, field)
		new_value = _safe_float(raw_value, old_value or 0.0)
		setattr(cierre, field, new_value)
		locked_fields.add(field)
		_audit(field, old_value, new_value)
		if field == 'venta_sistema':
			cierre.venta_sistema_guardada = True

	if 'observaciones' in fields:
		old_value = cierre.observaciones or ''
		new_value = str(fields.get('observaciones') or '')
		if old_value != new_value:
			cierre.observaciones = new_value
			_audit('observaciones', old_value, new_value)

	audit_enabled = bool(
		cierre.venta_sistema_guardada
		and locked_fields.intersection({'pos_tarjetas', 'yape', 'plin', 'efectivo'})
	)

	if 'efectivo_entregado' in fields and (('efectivo_entregado' not in locked_fields) or is_admin_general):
		raw_value = fields.get('efectivo_entregado')
		if raw_value not in ('', None) and (audit_enabled or is_admin_general):
			old_value = cierre.efectivo_entregado
			cierre.efectivo_entregado = _safe_float(raw_value, old_value or 0.0)
			cierre.efectivo_entregado_guardado = True
			locked_fields.add('efectivo_entregado')
			_audit('efectivo_entregado', old_value, cierre.efectivo_entregado)

	if 'efectivo_dejado_caja_real' in fields and (('efectivo_dejado_caja_real' not in locked_fields) or is_admin_general):
		raw_value = fields.get('efectivo_dejado_caja_real')
		if raw_value not in ('', None) and (cierre.efectivo_entregado_guardado or is_admin_general):
			recommended = (cierre.efectivo or 0.0) - (cierre.efectivo_entregado or 0.0)
			old_value = cierre.efectivo_dejado_caja_real
			cierre.efectivo_dejado_caja_recomendado = recommended
			cierre.efectivo_dejado_caja_real = _safe_float(raw_value, old_value or 0.0)
			cierre.diferencia_efectivo_dejado = cierre.efectivo_dejado_caja_real - recommended
			cierre.efectivo_dejado_guardado = True
			locked_fields.add('efectivo_dejado_caja_real')
			_audit('efectivo_dejado_caja_real', old_value, cierre.efectivo_dejado_caja_real)

	if 'gastos' in fields:
		incoming = fields.get('gastos') or []
		try:
			current_gastos = json.loads(cierre.gastos_json or '[]')
		except (TypeError, ValueError):
			current_gastos = []
		current_gastos, _ = _normalizar_gastos_arqueo(current_gastos)

		if is_admin_general:
			new_gastos_list = []
			for idx, item in enumerate(incoming):
				monto = _safe_float(item.get('monto'), 0.0)
				nombre = str(item.get('nombre') or '').strip()
				tipo = str(item.get('tipo') or 'Otros').strip()
				if monto > 0 or nombre:
					item_id = str(item.get('id') or f'gasto-{uuid.uuid4().hex}')
					new_gastos_list.append({
						'id': item_id,
						'tipo': tipo,
						'nombre': nombre,
						'monto': monto,
						'bloqueado': True,
					})
			cierre.gastos_json = json.dumps(new_gastos_list, ensure_ascii=True)
			gastos_antes, gastos_despues = _cambios_gastos_arqueo(current_gastos, new_gastos_list)
			if gastos_antes or gastos_despues:
				_audit('gastos', gastos_antes, gastos_despues)
		else:
			by_id = {str(item.get('id')): dict(item) for item in current_gastos if item.get('id') is not None}
			for idx, item in enumerate(incoming):
				monto = _safe_float(item.get('monto'), 0.0)
				nombre = str(item.get('nombre') or '').strip()
				tipo = str(item.get('tipo') or 'Otros').strip()
				if monto <= 0 and not nombre:
					continue
				item_id = str(item.get('id') or f'gasto-{uuid.uuid4().hex[:8]}')
				if item_id in by_id:
					existing = by_id[item_id]
					if existing.get('bloqueado'):
						# Sólo la propina puede aumentar y únicamente antes del cierre definitivo.
						if (existing.get('tipo') == 'Propina' or tipo == 'Propina') and not cierre.efectivo_dejado_guardado:
							prev_monto = _safe_float(existing.get('monto'), 0.0)
							if monto >= prev_monto:
								existing['monto'] = monto
								if nombre:
									existing['nombre'] = nombre
					else:
						existing['tipo'] = tipo
						existing['nombre'] = nombre
						existing['monto'] = monto
						existing['bloqueado'] = True
				else:
					by_id[item_id] = {
						'id': item_id,
						'tipo': tipo,
						'nombre': nombre,
						'monto': monto,
						'bloqueado': True,
					}
			cierre.gastos_json = json.dumps(list(by_id.values()), ensure_ascii=True)
			gastos_finales = list(by_id.values())
			gastos_antes, gastos_despues = _cambios_gastos_arqueo(current_gastos, gastos_finales)
			if gastos_antes or gastos_despues:
				_audit('gastos', gastos_antes, gastos_despues)

	sede_obj = Sede.query.get(target_sede_id)
	expected_base = _safe_float(sede_obj.monto_inicial_base_esperado if sede_obj else 0.0)
	cierre.efectivo_a_entregar = (cierre.efectivo or 0.0) - expected_base
	res_calc = _calc_cierre_operativo(
		cierre.monto_inicial or 0.0,
		cierre.pos_tarjetas or 0.0,
		cierre.yape or 0.0,
		cierre.plin or 0.0,
		cierre.efectivo or 0.0,
		cierre.venta_sistema or 0.0,
		json.loads(cierre.gastos_json or '[]')
	)
	cierre.monto_final = res_calc['subtotal']
	cierre.campos_bloqueados_json = json.dumps(sorted(locked_fields))
	db.session.commit()

	gastos_final = json.loads(cierre.gastos_json or '[]')
	logs = _historial_arqueo_por_alcance(cierre)
	logs_payload = [
		{
			'id': log.id_historial,
			'fecha_hora': log.fecha_hora.strftime('%Y-%m-%d %H:%M:%S') if log.fecha_hora else '',
			'usuario_id': log.usuario_id or '',
			'tipo_evento': log.tipo_evento or 'GUARDADO_MANUAL',
			'campo': log.campo_o_seccion_afectada or log.accion or '',
			'valor_anterior': log.valor_anterior or '',
			'valor_nuevo': log.valor_nuevo or '',
		}
		for log in logs
	]
	audit_enabled = bool(
		cierre.venta_sistema_guardada
		and locked_fields.intersection({'pos_tarjetas', 'yape', 'plin', 'efectivo'})
	)
	return {
		'ok': True,
		'locked_fields': sorted(locked_fields),
		'gastos': gastos_final,
		'venta_sistema_guardada': bool(cierre.venta_sistema_guardada),
		'audit_enabled': audit_enabled,
		'efectivo_entregado_guardado': bool(cierre.efectivo_entregado_guardado),
		'efectivo_dejado_guardado': bool(cierre.efectivo_dejado_guardado),
		'closure_complete': bool(cierre.efectivo_dejado_guardado),
		'resumen': res_calc,
		'logs': logs_payload,
		'message': 'Guardado correctamente.',
	}


def _audit_arqueo(arqueo, accion, valor_anterior, valor_nuevo):
	try:
		actor = current_user.id_usuario if current_user and not current_user.is_anonymous else None
	except Exception:
		actor = None
	return _audit_arqueo_event_with_user(arqueo, 'GUARDADO_MANUAL', accion, valor_anterior, valor_nuevo, actor)


def _audit_arqueo_event(arqueo, tipo_evento, campo, valor_anterior, valor_nuevo):
	try:
		actor = current_user.id_usuario if current_user and not current_user.is_anonymous else None
	except Exception:
		actor = None
	_audit_arqueo_event_with_user(arqueo, tipo_evento, campo, valor_anterior, valor_nuevo, actor)


def _audit_arqueo_event_with_user(arqueo, tipo_evento, campo, valor_anterior, valor_nuevo, actor_id):
	db.session.add(
		ArqueoCajaHistorial(
			id_arqueo=arqueo.id_arqueo,
			usuario_id=actor_id,
			accion=campo,
			tipo_evento=tipo_evento,
			campo_o_seccion_afectada=campo,
			valor_anterior=json.dumps(valor_anterior, ensure_ascii=True, default=str),
			valor_nuevo=json.dumps(valor_nuevo, ensure_ascii=True, default=str),
		)
	)


def _historial_arqueo_por_alcance(cierre, limite=50):
	"""Reúne la auditoría del mismo cierre operativo, incluso con IDs históricos duplicados."""
	if not cierre:
		return []
	return ArqueoCajaHistorial.query.join(
		ArqueoCaja, ArqueoCaja.id_arqueo == ArqueoCajaHistorial.id_arqueo
	).filter(
		ArqueoCaja.fecha == cierre.fecha,
		ArqueoCaja.id_sede == cierre.id_sede,
		ArqueoCaja.id_turno == cierre.id_turno,
	).order_by(ArqueoCajaHistorial.fecha_hora.desc()).limit(limite).all()


def _calc_cierre_operativo(monto_inicial, pos_tarjetas, yape, plin, efectivo, venta_sistema, gastos):
	total_ingresos = pos_tarjetas + yape + plin + efectivo
	gastos_totales = sum(_safe_float(item.get('monto'), 0.0) for item in (gastos or []))
	subtotal = total_ingresos + gastos_totales
	diferencia = (subtotal - monto_inicial) - venta_sistema
	estado_diferencia = 'Cuadre exacto'
	if diferencia > 0:
		estado_diferencia = 'Sobrante'
	elif diferencia < 0:
		estado_diferencia = 'Faltante'
	return {
		'total_ingresos': total_ingresos,
		'gastos_totales': gastos_totales,
		'subtotal': subtotal,
		'diferencia': diferencia,
		'estado_diferencia': estado_diferencia,
	}


def _normalize_header(text):
	if text is None:
		return ''
	return str(text).strip().lower().replace(' ', '_')


def _normalize_area(value):
	return str(value or '').strip().lower()


def _normalize_subarea(area, value):
	return str(value or '').strip().lower()


def _get_subareas_for_area(area_name):
	area_name = _normalize_area(area_name)
	if not area_name:
		return []
	area = Area.query.filter(db.func.lower(Area.nombre_area) == area_name).first()
	if not area:
		return []
	return [subarea.nombre_subarea for subarea in Subarea.query.filter_by(id_area=area.id_area).order_by(Subarea.nombre_subarea.asc()).all()]


def _checklist_base_query(user, selected_date=None):
	if not user.id_sede or not user.id_turno:
		return ChecklistPedido.query.filter(text('1=0'))
	query = ChecklistPedido.query.filter(
		ChecklistPedido.id_sede == user.id_sede,
		ChecklistPedido.id_turno == user.id_turno,
	)
	if selected_date is not None:
		query = query.filter(db.func.date(ChecklistPedido.fecha) == selected_date.strftime('%Y-%m-%d'))
	return query


def _get_active_checklist(user, selected_date=None):
	return _checklist_base_query(user, selected_date).filter(
		ChecklistPedido.estado_general.in_(['Borrador', 'Pendiente', 'Enviado'])
	).order_by(ChecklistPedido.id_pedido.desc()).first()


def _get_visible_checklist(user, selected_date=None):
	active = _get_active_checklist(user, selected_date)
	if active:
		return active
	return _checklist_base_query(user, selected_date).order_by(ChecklistPedido.id_pedido.desc()).first()


def _get_checklist_items(pedido, user=None, include_all=False, target_user_id='', target_area=''):
	if not pedido:
		return []
	query = db.session.query(DetallePedido, Producto).join(
		Producto, Producto.id_producto == DetallePedido.id_producto
	).filter(
		DetallePedido.id_pedido == pedido.id_pedido
	)
	if not include_all and user is not None:
		effective_user_id = target_user_id or user.id_usuario
		query = query.filter(
			or_(
				DetallePedido.id_usuario == effective_user_id,
				DetallePedido.id_usuario.is_(None),
			)
		)
	if include_all and target_user_id:
		query = query.filter(DetallePedido.id_usuario == target_user_id)
	if include_all and target_area:
		query = query.filter(db.func.lower(Producto.area) == target_area.lower())
	# No filtrar los items del pedido por inventario central: los items ya
	# existentes en `DetallePedido` deben seguir mostrándose en "Lista" y
	# "Editar lista" aunque el inventario central cambie. Solo el catálogo
	# (fuente para agregar items) se obtiene desde el inventario central.
	rows = query.order_by(Producto.nombre_producto.asc()).all()

	# Adjuntar nombre legible de categoria (`categoria_display`) a cada Producto
	# usando la misma lógica que en _get_checklist_catalog para asegurar
	# consistencia entre catálogo e items del checklist.
	if rows:
		all_categorias = {c.nombre_categoria: c.nombre_categoria for c in Categoria.query.all()}
		for detalle, producto in rows:
			cat_display = ''
			if producto and getattr(producto, 'id_area', None):
				found = next((c for c in all_categorias.keys() if c.lower() == (producto.id_area or '').lower()), None)
				if found:
					cat_display = all_categorias[found]
				else:
					try:
						cid = int(producto.id_area)
						categoria_obj = Categoria.query.get(cid)
						if categoria_obj:
							cat_display = categoria_obj.nombre_categoria
					except Exception:
						pass
			if not cat_display:
				cat_display = producto.id_area or '' if producto else ''
			if producto:
				setattr(producto, 'categoria_display', cat_display)

		# Ordenar filas para agrupar por categoria_display y luego por nombre de producto
		rows.sort(key=lambda pair: ((pair[1].categoria_display or '').lower(), (pair[1].nombre_producto or '').lower()))

	return rows


def _get_checklist_catalog(user, q='', id_sede=None):
	# Usar solo el inventario central (Almacen) como fuente única de productos para
	# el catálogo del checklist, evitando duplicados cuando hay inventarios por sede.
	almacen = Sede.query.filter(db.func.lower(Sede.nombre_sede) == 'almacen').first()
	if almacen:
		# Obtener los id_producto presentes en el inventario del almacen
		prod_ids = [r for (r,) in InventarioSede.query.with_entities(InventarioSede.id_producto).filter_by(id_sede=almacen.id_sede).all()]
		query = Producto.query.filter(Producto.id_producto.in_(prod_ids))
	else:
		# Fallback: unir a InventarioSede sin filtrar por sede (antiguo comportamiento)
		query = Producto.query.join(
			InventarioSede,
			InventarioSede.id_producto == Producto.id_producto,
		).distinct()
	if q:
		like_q = f"%{q}%"
		query = query.filter(
			or_(
				Producto.nombre_producto.ilike(like_q),
				Producto.id_producto.ilike(like_q),
				Producto.id_area.ilike(like_q),
				Producto.area.ilike(like_q),
				Producto.subarea.ilike(like_q),
			)
		)

	preferred_area = _preferred_area_for_user(user)
	productos = query.all()

	# Resolver nombre de categoría legible para cada producto y adjuntarlo como
	# atributo `categoria_display` para uso en plantillas. Producto.id_area puede
	# contener nombre de categoría o un identificador; intentamos coincidir con
	# la tabla Categoria por nombre o por id (cuando sea numérico).
	all_categorias = {c.nombre_categoria: c.nombre_categoria for c in Categoria.query.all()}
	for p in productos:
		cat_display = ''
		if p.id_area:
			# intento directo por nombre (case-insensitive)
			found = next((c for c in all_categorias.keys() if c.lower() == (p.id_area or '').lower()), None)
			if found:
				cat_display = all_categorias[found]
			else:
				# si id_area tiene formato numérico, buscar por id_categoria
				try:
					cid = int(p.id_area)
					categoria_obj = Categoria.query.get(cid)
					if categoria_obj:
						cat_display = categoria_obj.nombre_categoria
				except Exception:
					pass
		if not cat_display:
			cat_display = p.id_area or ''
		setattr(p, 'categoria_display', cat_display)
	# Ordenar por categoria legible y luego por nombre de producto para que en la
	# interfaz las categorias aparezcan agrupadas.
	def _sort_key(p):
		return ((p.categoria_display or '').lower(), (p.nombre_producto or '').lower())

	if not preferred_area:
		return sorted(productos, key=_sort_key)
	# Si hay preferencia de area, mantenerla como prioridad secundaria
	return sorted(
		productos,
		key=lambda p: (0 if (p.area or '').lower() == preferred_area else 1, _sort_key(p)),
	)


def _preferred_area_for_user(user):
	return _preferred_area_for_role_name(user.rol_nombre)


def _preferred_area_for_role_name(role_name):
	if role_name == 'cocinero':
		return 'cocina'
	if role_name == 'admin_sala':
		return 'sala'
	return ''


def _checklist_scope_users(user):
	allowed_roles = {'cocinero', 'admin_sala'}
	preferred_area = _preferred_area_for_user(user)
	users = Usuario.query.join(Rol, Rol.id_rol == Usuario.id_rol).filter(
		Usuario.id_sede == user.id_sede,
		Usuario.id_turno == user.id_turno,
		Rol.nombre_rol.in_(allowed_roles),
	).all()
	if preferred_area:
		users = [scope_user for scope_user in users if _preferred_area_for_user(scope_user) == preferred_area]
	if not users:
		return [user]
	return users


def _template_scope_query(user):
	return PlantillaChecklistItem.query.filter_by(
		id_usuario=user.id_usuario,
		id_sede=user.id_sede,
		id_turno=user.id_turno,
		area=_preferred_area_for_user(user),
	)


def _get_template_product_ids(user):
	return _get_template_product_ids_for_user(user, user.id_usuario)


def _get_template_product_ids_for_user(user, target_user_id):
	target_user_id = (target_user_id or '').strip()
	if not target_user_id:
		return set()
	return {
		id_producto
		for (id_producto,) in PlantillaChecklistItem.query.join(
			Producto, Producto.id_producto == PlantillaChecklistItem.id_producto
		).filter(
			PlantillaChecklistItem.id_usuario == target_user_id,
			PlantillaChecklistItem.id_sede == user.id_sede,
			PlantillaChecklistItem.id_turno == user.id_turno,
			PlantillaChecklistItem.area == _preferred_area_for_user(user),
		).filter(
			or_(Producto.estado.is_(None), Producto.estado == '', Producto.estado == 'Activo')
		).with_entities(PlantillaChecklistItem.id_producto).all()
	}


def _build_template_export_payload(user, target_user_id=None):
	target_user_id = (target_user_id or user.id_usuario or '').strip()
	target_user = db.session.get(Usuario, target_user_id) if target_user_id else user
	items = (
		db.session.query(PlantillaChecklistItem, Producto)
		.join(Producto, Producto.id_producto == PlantillaChecklistItem.id_producto)
		.filter(
			PlantillaChecklistItem.id_usuario == target_user_id,
			PlantillaChecklistItem.id_sede == user.id_sede,
			PlantillaChecklistItem.id_turno == user.id_turno,
			PlantillaChecklistItem.area == _preferred_area_for_user(user),
		)
		.order_by(Producto.nombre_producto.asc())
		.all()
	)
	return {
		'meta': {
			'tipo': 'plantilla_checklist',
			'usuario': target_user.username if target_user else '',
			'id_usuario': target_user_id,
			'id_sede': user.id_sede,
			'id_turno': user.id_turno,
			'area': _preferred_area_for_user(user),
			'exportado_en': datetime.utcnow().isoformat(timespec='seconds') + 'Z',
		},
		'productos': [
			{
				'id_producto': producto.id_producto,
				'nombre_producto': producto.nombre_producto,
				'unidad': producto.unidad or '',
				'area': producto.area or '',
			}
			for _, producto in items
		],
	}


def _normalize_template_import_payload(payload):
	if isinstance(payload, dict):
		products = payload.get('productos') or payload.get('items') or payload.get('products') or []
	else:
		products = payload or []

	product_ids = []
	for item in products:
		if isinstance(item, str):
			candidate = item.strip()
		elif isinstance(item, dict):
			candidate = str(item.get('id_producto') or item.get('producto_id') or item.get('id') or '').strip()
		else:
			candidate = ''
		if candidate:
			product_ids.append(candidate)

	seen = set()
	ordered_unique = []
	for id_producto in product_ids:
		if id_producto in seen:
			continue
		seen.add(id_producto)
		ordered_unique.append(id_producto)
	return ordered_unique


def _replace_template_from_import(user, target_user_id, payload):
	product_ids = _normalize_template_import_payload(payload)
	if not product_ids:
		return {'importados': 0, 'omitidos': 0, 'total': 0}

	target_user = db.session.get(Usuario, target_user_id) if target_user_id else None
	if not target_user:
		target_user = user

	allowed_ids = set(
		row[0]
		for row in db.session.query(Producto.id_producto).filter(
			Producto.id_producto.in_(product_ids),
			or_(Producto.estado.is_(None), Producto.estado == '', Producto.estado == 'Activo'),
		).all()
	)
	PlantillaChecklistItem.query.filter_by(
		id_usuario=target_user.id_usuario,
		id_sede=target_user.id_sede,
		id_turno=target_user.id_turno,
		area=_preferred_area_for_user(target_user),
	).delete(synchronize_session=False)

	importados = 0
	omitidos = 0
	for id_producto in product_ids:
		if id_producto not in allowed_ids:
			omitidos += 1
			continue
		db.session.add(
			PlantillaChecklistItem(
				id_usuario=target_user.id_usuario,
				id_sede=target_user.id_sede,
				id_turno=target_user.id_turno,
				area=_preferred_area_for_user(target_user),
				id_producto=id_producto,
			)
		)
		importados += 1

	return {'importados': importados, 'omitidos': omitidos, 'total': len(product_ids)}


def _sync_checklist_items_with_template(checklist, template_product_ids, user_id):
	if not checklist or checklist.estado_general not in {'Borrador', 'Pendiente'}:
		return

	user_id = (user_id or '').strip()
	if not user_id:
		return

	existing_items = DetallePedido.query.filter(
		DetallePedido.id_pedido == checklist.id_pedido,
		DetallePedido.id_usuario == user_id,
	).all()

	existing_by_product = {item.id_producto: item for item in existing_items}
	for id_producto in template_product_ids:
		if id_producto in existing_by_product:
			continue
		db.session.add(
			DetallePedido(
				id_pedido=checklist.id_pedido,
				id_usuario=user_id,
				id_producto=id_producto,
				cantidad_pedida=0.0,
				estado_sede='Pendiente',
			)
		)

	for item in existing_items:
		if item.id_producto in template_product_ids:
			continue
		if item.estado_sede == 'Recibido' or _safe_float(item.cantidad_entregada, 0.0) > 0:
			continue
		db.session.delete(item)


def _sync_open_checklists_with_template(user, selected_date):
	template_product_ids = _get_template_product_ids(user)

	open_checklists = ChecklistPedido.query.filter(
		ChecklistPedido.id_sede == user.id_sede,
		ChecklistPedido.id_turno == user.id_turno,
		db.func.date(ChecklistPedido.fecha) >= selected_date.strftime('%Y-%m-%d'),
		ChecklistPedido.estado_general.in_(['Borrador', 'Pendiente']),
	).order_by(ChecklistPedido.fecha.asc(), ChecklistPedido.id_pedido.asc()).all()

	for checklist in open_checklists:
		_sync_checklist_items_with_template(checklist, template_product_ids, user.id_usuario)


def _build_checklist_from_template_if_needed(user, selected_date):
	today = datetime.now().date()
	if selected_date < today:
		return None

	template_product_ids = _get_template_product_ids(user)
	if not template_product_ids:
		return None

	current = _checklist_base_query(user, selected_date).filter(
		ChecklistPedido.estado_general.in_(['Borrador', 'Pendiente'])
	).order_by(ChecklistPedido.id_pedido.desc()).first()
	if current:
		_sync_checklist_items_with_template(current, template_product_ids, user.id_usuario)
		return current

	locked = _checklist_base_query(user, selected_date).filter(
		ChecklistPedido.estado_general.in_(['Enviado', 'Finalizado'])
	).first()
	if locked:
		return None

	checklist = ChecklistPedido(
		id_sede=user.id_sede,
		id_turno=user.id_turno,
		id_usuario=user.id_usuario,
		fecha=datetime.combine(selected_date, datetime.min.time()),
		estado_general='Borrador',
	)
	db.session.add(checklist)
	db.session.flush()
	for id_producto in sorted(template_product_ids):
		db.session.add(
			DetallePedido(
				id_pedido=checklist.id_pedido,
				id_usuario=user.id_usuario,
				id_producto=id_producto,
				cantidad_pedida=0.0,
				estado_sede='Pendiente',
			)
		)
	return checklist


def _complete_checklist_if_all_received(pedido):
	if not pedido or pedido.estado_general != 'Enviado':
		return
	pending = DetallePedido.query.filter(
		DetallePedido.id_pedido == pedido.id_pedido,
		DetallePedido.cantidad_entregada > 0,
		DetallePedido.estado_sede != 'Recibido',
	).count()
	if pending == 0:
		pedido.estado_general = 'Finalizado'


def _ensure_inventory_schema(app):
	inspector = inspect(db.engine)
	columns = {column['name'] for column in inspector.get_columns('productos')}
	if 'unidad' not in columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE productos ADD COLUMN unidad VARCHAR(50)'))
	if 'area' not in columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE productos ADD COLUMN area VARCHAR(20)'))
	if 'costo_unitario' not in columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE productos ADD COLUMN costo_unitario FLOAT DEFAULT 0'))

	detalle_columns = {column['name'] for column in inspector.get_columns('detalle_pedido')}
	if 'id_usuario' not in detalle_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE detalle_pedido ADD COLUMN id_usuario VARCHAR(50)'))

	arqueo_columns = {column['name'] for column in inspector.get_columns('arqueo_caja')}
	if 'pos_tarjetas' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN pos_tarjetas FLOAT DEFAULT 0'))
	if 'yape' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN yape FLOAT DEFAULT 0'))
	if 'plin' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN plin FLOAT DEFAULT 0'))
	if 'efectivo' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN efectivo FLOAT DEFAULT 0'))
	if 'efectivo_a_entregar' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN efectivo_a_entregar FLOAT DEFAULT 0'))
	if 'venta_sistema' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN venta_sistema FLOAT DEFAULT 0'))
	if 'gastos_json' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text("ALTER TABLE arqueo_caja ADD COLUMN gastos_json TEXT DEFAULT '[]'"))
	if 'efectivo_entregado' not in arqueo_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE arqueo_caja ADD COLUMN efectivo_entregado FLOAT DEFAULT 0'))
	for column_name, column_type in (
		('efectivo_dejado_caja_recomendado', 'FLOAT DEFAULT 0'),
		('efectivo_dejado_caja_real', 'FLOAT DEFAULT 0'),
		('diferencia_efectivo_dejado', 'FLOAT DEFAULT 0'),
		('seccion_1_guardada', 'BOOLEAN DEFAULT FALSE'),
		('efectivo_entregado_guardado', 'BOOLEAN DEFAULT FALSE'),
		('efectivo_dejado_guardado', 'BOOLEAN DEFAULT FALSE'),
		('campos_bloqueados_json', "TEXT DEFAULT '[]'"),
		('venta_sistema_guardada', 'BOOLEAN DEFAULT FALSE'),
	):
		if column_name not in arqueo_columns:
			with db.engine.begin() as connection:
				connection.execute(text(f'ALTER TABLE arqueo_caja ADD COLUMN {column_name} {column_type}'))

	historial_columns = {column['name'] for column in inspector.get_columns('arqueo_caja_historial')}
	for column_name, column_type in (
		('tipo_evento', "VARCHAR(30) DEFAULT 'GUARDADO_MANUAL'"),
		('campo_o_seccion_afectada', 'VARCHAR(120)'),
	):
		if column_name not in historial_columns:
			with db.engine.begin() as connection:
				connection.execute(text(f'ALTER TABLE arqueo_caja_historial ADD COLUMN {column_name} {column_type}'))

	# Asegurar columna en tabla sedes para monto inicial base esperado
	sedes_columns = {column['name'] for column in inspector.get_columns('sedes')}
	if 'monto_inicial_base_esperado' not in sedes_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE sedes ADD COLUMN monto_inicial_base_esperado FLOAT DEFAULT 0'))

	usuario_columns = {column['name'] for column in inspector.get_columns('usuarios')}
	if 'dni' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN dni VARCHAR(20)'))
	if 'fecha_nacimiento' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN fecha_nacimiento DATE'))
	if 'email' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN email VARCHAR(120)'))
	if 'telefono' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN telefono VARCHAR(30)'))
	if 'direccion' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN direccion VARCHAR(180)'))
	if 'bio' not in usuario_columns:
		with db.engine.begin() as connection:
			connection.execute(text('ALTER TABLE usuarios ADD COLUMN bio VARCHAR(240)'))

	with db.engine.begin() as connection:
		if not inspect(db.engine).has_table('recordatorios_cierre'):
			RecordatorioCierre.__table__.create(bind=db.engine)
		connection.execute(text("UPDATE arqueo_caja SET venta_sistema_guardada = TRUE WHERE venta_sistema IS NOT NULL AND venta_sistema != 0 AND venta_sistema_guardada = FALSE"))
		connection.execute(
			text(
				"""
				UPDATE detalle_pedido
				SET id_usuario = (
					SELECT checklist_pedidos.id_usuario
					FROM checklist_pedidos
					WHERE checklist_pedidos.id_pedido = detalle_pedido.id_pedido
				)
				WHERE id_usuario IS NULL
				"""
			)
		)

	if app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
		with db.engine.begin() as connection:
			connection.execute(text("UPDATE productos SET unidad = COALESCE(unidad, 'unidad') WHERE unidad IS NULL OR unidad = ''"))
			connection.execute(text("UPDATE productos SET area = COALESCE(NULLIF(area, ''), 'cocina')"))
			connection.execute(text("UPDATE productos SET subarea = 'cocina_caliente' WHERE area = 'cocina' AND (subarea IS NULL OR subarea = '')"))
			connection.execute(text("UPDATE productos SET subarea = 'sala' WHERE area = 'sala' AND (subarea IS NULL OR subarea = '')"))
			connection.execute(text("UPDATE arqueo_caja SET gastos_json = '[]' WHERE gastos_json IS NULL OR gastos_json = ''"))

	# Consolidar duplicados históricos en arqueo_caja y asegurar índice único
	try:
		with db.engine.begin() as connection:
			dup_rows = connection.execute(text("""
				SELECT id_sede, id_turno, fecha, COUNT(*) as cnt
				FROM arqueo_caja
				WHERE id_sede IS NOT NULL AND id_turno IS NOT NULL AND fecha IS NOT NULL
				GROUP BY id_sede, id_turno, fecha
				HAVING COUNT(*) > 1
			""")).fetchall()

			for dup in dup_rows:
				s_id, t_id, f_val = dup[0], dup[1], dup[2]
				records = connection.execute(
					text("SELECT id_arqueo FROM arqueo_caja WHERE id_sede = :s AND id_turno = :t AND fecha = :f ORDER BY id_arqueo DESC"),
					{'s': s_id, 't': t_id, 'f': f_val}
				).fetchall()
				if records and len(records) > 1:
					primary_id = records[0][0]
					dup_ids = [r[0] for r in records[1:]]
					for d_id in dup_ids:
						connection.execute(
							text("UPDATE arqueo_caja_historial SET id_arqueo = :prim WHERE id_arqueo = :dup"),
							{'prim': primary_id, 'dup': d_id}
						)
						connection.execute(
							text("DELETE FROM arqueo_caja WHERE id_arqueo = :dup"),
							{'dup': d_id}
						)

			connection.execute(text('CREATE UNIQUE INDEX IF NOT EXISTS uq_arqueo_caja_sede_turno_fecha ON arqueo_caja (id_sede, id_turno, fecha)'))
	except Exception as err:
		print(f'Aviso al asegurar índice uq_arqueo_caja_sede_turno_fecha: {err}')


def create_app():
	app = Flask(__name__)

	# Render puede entregar postgres://; SQLAlchemy espera postgresql://
	database_url = os.environ.get('DATABASE_URL')
	if database_url and database_url.startswith('postgres://'):
		database_url = database_url.replace('postgres://', 'postgresql://', 1)

	# Fallback local para desarrollo cuando no existe DATABASE_URL
	app.config['SQLALCHEMY_DATABASE_URI'] = database_url or 'sqlite:///mi_app.db'
	app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
	# Evita que el navegador conserve plantillas y recursos durante el desarrollo.
	app.config['TEMPLATES_AUTO_RELOAD'] = True
	app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

	secret_key = os.environ.get('SECRET_KEY')
	is_production = bool(
		os.environ.get('RAILWAY_ENVIRONMENT')
		or os.environ.get('RENDER')
		or os.environ.get('FLASK_ENV') == 'production'
		or os.environ.get('PRODUCTION')
	)
	if not secret_key:
		if is_production:
			raise RuntimeError('SECRET_KEY environment variable is required in production!')
		secret_key = 'dev_key_punto29_' + secrets.token_hex(16)
	app.config['SECRET_KEY'] = secret_key

	app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=90)
	app.config['REMEMBER_COOKIE_REFRESH_EACH_REQUEST'] = True
	app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=90)
	app.config['SESSION_COOKIE_HTTPONLY'] = True
	app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
	app.jinja_env.filters['peru_datetime'] = _format_peru_datetime

	def _get_csrf_token():
		if '_csrf_token' not in session:
			session['_csrf_token'] = secrets.token_hex(32)
		return session['_csrf_token']

	app.jinja_env.globals['csrf_token'] = _get_csrf_token

	@app.before_request
	def csrf_protect():
		if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
			token = (
				request.headers.get('X-CSRFToken')
				or request.headers.get('X-CSRF-Token')
				or request.form.get('csrf_token')
			)
			if not token and request.is_json:
				payload = request.get_json(silent=True) or {}
				if isinstance(payload, dict):
					token = payload.get('csrf_token')
			expected = session.get('_csrf_token')
			if not expected or not token or not hmac.compare_digest(str(token), str(expected)):
				if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
					return jsonify({'ok': False, 'error': 'csrf_error', 'message': 'Token CSRF inválido o expirado. Recarga la página.'}), 403
				flash('Sesión o token de seguridad inválido. Por favor intenta de nuevo.', 'error')
				return redirect(request.referrer or url_for('dashboard'))

	@app.after_request
	def disable_browser_cache(response):
		response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
		response.headers['Pragma'] = 'no-cache'
		response.headers['Expires'] = '0'
		return response

	db.init_app(app)
	login_manager.init_app(app)

	with app.app_context():
		db.create_all()
		_ensure_inventory_schema(app)
		_seed_catalogs()
		print('Base de datos vinculada y tablas creadas.')

	@app.route('/')
	def index():
		if current_user.is_authenticated:
			return redirect(url_for('dashboard'))
		return redirect(url_for('login'))

	@app.route('/set-app-date', methods=['POST'])
	@login_required
	def set_app_date():
		selected_date = request.form.get('app_date', '').strip()
		try:
			datetime.strptime(selected_date, '%Y-%m-%d')
		except ValueError:
			flash('Fecha invalida.', 'error')
			return redirect(request.referrer or url_for('dashboard'))

		session['app_date'] = selected_date
		return redirect(request.referrer or url_for('dashboard'))

	@app.route('/login', methods=['GET', 'POST'])
	def login():
		if current_user.is_authenticated:
			return redirect(url_for('dashboard'))

		if request.method == 'POST':
			username = request.form.get('username', '').strip()
			password = request.form.get('password', '')
			user = Usuario.query.filter_by(username=username).first()

			is_valid = False
			if user:
				is_valid = check_password_hash(user.password_hash, password) or user.password_hash == password

			if is_valid:
				session.permanent = True
				login_user(user, remember=True)
				return redirect(url_for('dashboard'))

			flash('Usuario o contrasena invalidos.', 'error')

		return render_template('login.html')

	@app.route('/logout')
	@login_required
	def logout():
		session.pop('app_date', None)
		logout_user()
		return redirect(url_for('login'))

	@app.route('/dashboard')
	@login_required
	def dashboard():
		selected_date = _get_selected_app_date()
		try:
			stats = _stats_for_user(current_user)
			alerts = _home_alerts_for_user(current_user, selected_date)
		except Exception:
			stats = {'productos': 0, 'movimientos': 0, 'pedidos': 0, 'arqueos': 0}
			alerts = {'stock_critico_count': 0, 'pedidos_pendientes_count': 0, 'missing_arqueo': False, 'subtitle': 'Resumen de tareas para hoy segun tu rol.', 'cards': [{'title': 'Sin pendientes', 'message': 'No se pudo calcular el resumen del dashboard. Reintenta.', 'state': 'ok', 'link': None}]}
		return render_template(
			'dashboard/home.html',
			allowed_views=_allowed_views(current_user),
			stats=stats,
			alerts=alerts,
		)

	@app.route('/inventario/dashboard')
	@login_required
	def inventario_dashboard():
		if not current_user.can_view('inventario'):
			return _forbidden_redirect()
		selected_date = _get_selected_app_date()
		metrics = _inventory_dashboard_metrics(current_user, selected_date)
		return render_template(
			'dashboard/inventario_dashboard.html',
			allowed_views=_allowed_views(current_user),
			selected_date=selected_date,
			metrics=metrics,
		)

	@app.route('/inventario/dashboard/export')
	@login_required
	def inventario_dashboard_export():
		if not current_user.can_view('inventario'):
			return _forbidden_redirect()

		selected_date = _get_selected_app_date()
		metrics = _inventory_dashboard_metrics(current_user, selected_date)
		openpyxl = importlib.import_module('openpyxl')
		wb = openpyxl.Workbook()

		ws_low = wb.active
		ws_low.title = 'Por acabarse'
		ws_low.append(['ID', 'Producto', 'Stock', 'Minimo', 'Sede'])
		for producto, inv, sede in metrics.get('por_acabarse', []):
			ws_low.append([
				producto.id_producto,
				producto.nombre_producto,
				inv.stock_actual if inv else 0,
				inv.punto_minimo if inv else 0,
				sede.nombre_sede if sede else '',
			])

		ws_zero = wb.create_sheet('Acabados')
		ws_zero.append(['ID', 'Producto', 'Stock', 'Sede'])
		for producto, inv, sede in metrics.get('acabados', []):
			ws_zero.append([
				producto.id_producto,
				producto.nombre_producto,
				inv.stock_actual if inv else 0,
				sede.nombre_sede if sede else '',
			])

		output = BytesIO()
		wb.save(output)
		output.seek(0)
		stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
		return send_file(
			output,
			as_attachment=True,
			download_name=f'inventario_alertas_{stamp}.xlsx',
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)

	@app.route('/perfil', methods=['GET', 'POST'])
	@login_required
	def perfil():
		if request.method == 'POST':
			current_user.dni = request.form.get('dni', '').strip() or None
			current_user.email = request.form.get('email', '').strip() or None
			current_user.telefono = request.form.get('telefono', '').strip() or None
			current_user.direccion = request.form.get('direccion', '').strip() or None
			current_user.bio = request.form.get('bio', '').strip() or None

			fecha_nacimiento_raw = request.form.get('fecha_nacimiento', '').strip()
			if fecha_nacimiento_raw:
				try:
					current_user.fecha_nacimiento = datetime.strptime(fecha_nacimiento_raw, '%Y-%m-%d').date()
				except ValueError:
					flash('Fecha de nacimiento invalida.', 'error')
					return redirect(url_for('perfil'))
			else:
				current_user.fecha_nacimiento = None

			password_actual = request.form.get('password_actual', '')
			nueva_password = request.form.get('nueva_password', '')
			confirm_password = request.form.get('confirm_password', '')
			if password_actual or nueva_password or confirm_password:
				password_ok = check_password_hash(current_user.password_hash, password_actual) or current_user.password_hash == password_actual
				if not password_ok:
					flash('La contraseña actual no coincide.', 'error')
					return redirect(url_for('perfil'))
				if not nueva_password:
					flash('Debes ingresar una nueva contraseña.', 'error')
					return redirect(url_for('perfil'))
				if nueva_password != confirm_password:
					flash('La confirmación de contraseña no coincide.', 'error')
					return redirect(url_for('perfil'))
				current_user.password_hash = generate_password_hash(nueva_password)

			db.session.commit()
			flash('Perfil actualizado correctamente.', 'ok')
			return redirect(url_for('perfil'))

		return render_template(
			'perfil.html',
			allowed_views=_allowed_views(current_user),
		)

	@app.route('/inventario', methods=['GET', 'POST'])
	@login_required
	def inventario():
		if not current_user.can_view('inventario'):
			return _forbidden_redirect()

		if request.method == 'POST':
			if not current_user.can_write('inventario', 'update'):
				return _forbidden_redirect()

			action = request.form.get('action', 'update_row')
			id_producto = request.form.get('id_producto', '').strip()
			target_sede = current_user.id_sede if current_user.rol_nombre != 'admin_general' else int(request.form.get('id_sede', current_user.id_sede))

			if action == 'upsert_product':
				if not id_producto:
					id_producto = _generate_product_id()

				producto = Producto.query.filter_by(id_producto=id_producto).first()
				if not producto:
					producto = Producto(id_producto=id_producto)
					db.session.add(producto)

				producto.nombre_producto = request.form.get('nombre_producto', '').strip()
				producto.id_area = request.form.get('id_area', '').strip()
				producto.area = _normalize_area(request.form.get('area', '')) or 'cocina'
				producto.subarea = _normalize_subarea(producto.area, request.form.get('subarea', ''))
				producto.unidad = request.form.get('unidad', '').strip()
				producto.costo_unitario = _safe_float(request.form.get('costo_unitario'), producto.costo_unitario or 0.0)
				producto.estado = request.form.get('estado', 'Activo').strip() or 'Activo'

				row = InventarioSede.query.filter_by(id_sede=target_sede, id_producto=id_producto).first()
				if not row:
					row = InventarioSede(id_sede=target_sede, id_producto=id_producto, stock_actual=0.0, punto_minimo=0.0)
					db.session.add(row)

				# Registrar movimiento si hay cambio de stock
				prev_stock = float(row.stock_actual or 0.0)
				new_stock = _safe_float(request.form.get('stock_actual'), prev_stock)
				row.stock_actual = new_stock
				row.punto_minimo = _safe_float(request.form.get('punto_minimo'), row.punto_minimo or 0.0)
				stock_delta = new_stock - prev_stock
				if abs(stock_delta) > 0.0001:
					tipo = 'ENTRADA' if stock_delta > 0 else 'SALIDA'
					motivo = 'Ajuste inventario (edicion producto)'
					db.session.add(
						MovimientoInventario(
							id_sede=target_sede,
							id_producto=id_producto,
							cantidad=abs(stock_delta),
							tipo=tipo,
							motivo=motivo,
							fecha=datetime.utcnow(),
							id_usuario=current_user.id_usuario,
						)
					)
				db.session.commit()
				flash('Producto guardado en inventario.', 'ok')

			elif action == 'delete_product':
				row = InventarioSede.query.filter_by(id_sede=target_sede, id_producto=id_producto).first()
				if row:
					db.session.delete(row)
					# Asegurarse que la eliminación pendiente se flushee antes de verificar filas restantes
					db.session.flush()
					PlantillaChecklistItem.query.filter_by(id_sede=target_sede, id_producto=id_producto).delete(synchronize_session=False)
					open_pedido_ids = [
						pedido_id
						for (pedido_id,) in db.session.query(ChecklistPedido.id_pedido).filter(
							ChecklistPedido.id_sede == target_sede,
							ChecklistPedido.estado_general.in_(['Borrador', 'Pendiente', 'Enviado']),
						).all()
					]
					if open_pedido_ids:
						DetallePedido.query.filter(
							DetallePedido.id_pedido.in_(open_pedido_ids),
							DetallePedido.id_producto == id_producto,
						).delete(synchronize_session=False)
					# Eliminar movimientos asociados antes de borrar el producto para evitar errores de FK
					MovimientoInventario.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
					# Si no quedan registros de inventario para este producto, entonces eliminar el producto
					if InventarioSede.query.filter_by(id_producto=id_producto).count() == 0:
						PlantillaChecklistItem.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
						DetallePedido.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
						producto = Producto.query.filter_by(id_producto=id_producto).first()
						if producto:
							db.session.delete(producto)
					try:
						db.session.commit()
					except IntegrityError as exc:
						db.session.rollback()
						flash(f'Error al eliminar producto: {exc}', 'error')
					else:
						flash('Producto eliminado del inventario.', 'ok')
				else:
					flash('No se encontro el producto en esa sede.', 'error')

			elif action == 'create_category':
				nombre_categoria = request.form.get('nombre_categoria', '').strip()
				if nombre_categoria and not Categoria.query.filter(db.func.lower(Categoria.nombre_categoria) == nombre_categoria.lower()).first():
					db.session.add(Categoria(nombre_categoria=nombre_categoria))
					db.session.commit()
					flash('Categoria creada.', 'ok')
				else:
					flash('La categoria ya existe o esta vacia.', 'error')

			elif action == 'delete_category':
				nombre_categoria = request.form.get('nombre_categoria', '').strip()
				if nombre_categoria and not Producto.query.filter(Producto.id_area == nombre_categoria).first():
					categoria = Categoria.query.filter_by(nombre_categoria=nombre_categoria).first()
					if categoria:
						db.session.delete(categoria)
						db.session.commit()
						flash('Categoria eliminada.', 'ok')
				else:
					flash('No se puede eliminar una categoria con productos asociados.', 'error')

			elif action == 'create_unit':
				nombre_unidad = request.form.get('nombre_unidad', '').strip()
				if nombre_unidad and not Unidad.query.filter(db.func.lower(Unidad.nombre_unidad) == nombre_unidad.lower()).first():
					db.session.add(Unidad(nombre_unidad=nombre_unidad))
					db.session.commit()
					flash('Unidad creada.', 'ok')
				else:
					flash('La unidad ya existe o esta vacia.', 'error')

			elif action == 'create_area':
				nombre_area = request.form.get('nombre_area', '').strip()
				if nombre_area and not Area.query.filter(db.func.lower(Area.nombre_area) == nombre_area.lower()).first():
					db.session.add(Area(nombre_area=nombre_area))
					db.session.commit()
					flash('Area creada.', 'ok')
				else:
					flash('El area ya existe o esta vacia.', 'error')

			elif action == 'delete_product':
				# Borrado en cascada: eliminar todo inventario y movimientos del producto a nivel global
				any_row = InventarioSede.query.filter_by(id_sede=target_sede, id_producto=id_producto).first()
				if any_row:
					# eliminar movimientos históricos del producto
					MovimientoInventario.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
					# eliminar todas las filas de inventario para este producto en todas las sedes
					InventarioSede.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
					# eliminar referencias en plantillas
					PlantillaChecklistItem.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
					# eliminar detalles de pedidos asociados
					DetallePedido.query.filter_by(id_producto=id_producto).delete(synchronize_session=False)
					# eliminar producto
					producto = Producto.query.filter_by(id_producto=id_producto).first()
					if producto:
						db.session.delete(producto)
					try:
						db.session.commit()
					except IntegrityError as exc:
						db.session.rollback()
						flash(f'Error al eliminar producto: {exc}', 'error')
					else:
						flash('Producto y sus registros relacionados fueron eliminados.', 'ok')
				else:
					flash('No se encontro el producto en esa sede.', 'error')

		q = request.args.get('q', '').strip()
		categoria = request.args.get('categoria', '').strip()
		area = _normalize_area(request.args.get('area', '').strip())
		subarea = request.args.get('subarea', '').strip()
		unidad = request.args.get('unidad', '').strip()

		inventario_rows = _inventory_query_for_user(current_user, q=q, categoria=categoria, subarea=subarea, unidad=unidad, area=area).all()
		categorias = Categoria.query.order_by(Categoria.nombre_categoria).all()
		unidades = Unidad.query.order_by(Unidad.nombre_unidad).all()
		areas = Area.query.order_by(Area.nombre_area).all()
		category_counts = {
			categoria_row.nombre_categoria: Producto.query.filter(Producto.id_area == categoria_row.nombre_categoria).count()
			for categoria_row in categorias
		}
		unit_counts = {
			unidad_row.nombre_unidad: Producto.query.filter(Producto.unidad == unidad_row.nombre_unidad).count()
			for unidad_row in unidades
		}
		subareas = _get_subareas_for_area(area or 'cocina')
		return render_template(
			'dashboard/inventario.html',
			allowed_views=_allowed_views(current_user),
			inventario_rows=inventario_rows,
			productos=Producto.query.order_by(Producto.nombre_producto).all(),
			categorias=categorias,
			unidades=unidades,
			areas=areas,
			category_counts=category_counts,
			unit_counts=unit_counts,
			subareas=subareas,
			selected_q=q,
			selected_categoria=categoria,
			selected_area=area,
			selected_subarea=subarea,
			selected_unidad=unidad,
			areas_subareas={area.nombre_area: _get_subareas_for_area(area.nombre_area) for area in areas},
			can_edit=current_user.can_write('inventario', 'update'),
		)

	@app.route('/inventario/export')
	@login_required
	def inventario_export():
		if not current_user.can_view('inventario'):
			return _forbidden_redirect()

		q = request.args.get('q', '').strip()
		categoria = request.args.get('categoria', '').strip()
		area = _normalize_area(request.args.get('area', '').strip())
		subarea = request.args.get('subarea', '').strip()
		unidad = request.args.get('unidad', '').strip()
		rows = _inventory_query_for_user(current_user, q=q, categoria=categoria, subarea=subarea, unidad=unidad, area=area).all()
		openpyxl = importlib.import_module('openpyxl')

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = 'Inventario'
		ws.append(['ID', 'Producto', 'Categoria', 'Area', 'Subarea', 'Unidad', 'Punto minimo', 'Stock central', 'Estado', 'Sede'])
		for producto, inv, sede in rows:
			ws.append([
				producto.id_producto,
				producto.nombre_producto,
				producto.id_area,
				producto.area,
				producto.subarea,
				producto.unidad or 'unidad',
				inv.punto_minimo if inv else 0,
				inv.stock_actual if inv else 0,
				producto.estado,
				sede.nombre_sede if sede else '',
			])

		output = BytesIO()
		wb.save(output)
		output.seek(0)
		stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
		return send_file(
			output,
			as_attachment=True,
			download_name=f'inventario_{stamp}.xlsx',
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)

	@app.route('/inventario/import', methods=['POST'])
	@login_required
	def inventario_import():
		if not current_user.can_write('inventario', 'update'):
			return _forbidden_redirect()

		file = request.files.get('excel_file')
		if not file or not file.filename.lower().endswith('.xlsx'):
			flash('Sube un archivo .xlsx valido.', 'error')
			return redirect(url_for('inventario'))

		try:
			openpyxl = importlib.import_module('openpyxl')
			wb = openpyxl.load_workbook(file)
		except ModuleNotFoundError:
			flash('No se encontro openpyxl en el entorno. Instala dependencias y vuelve a intentar.', 'error')
			return redirect(url_for('inventario'))
		except Exception as exc:
			flash(f'No se pudo leer el archivo Excel: {exc}', 'error')
			return redirect(url_for('inventario'))

		ws = wb.active
		rows = ws.iter_rows(values_only=True)
		headers = next(rows, None)
		if not headers:
			flash('El archivo no tiene encabezados.', 'error')
			return redirect(url_for('inventario'))

		header_map = {_normalize_header(name): idx for idx, name in enumerate(headers)}
		aliases = {
			'id': ['id', 'id_producto', 'codigo'],
			'producto': ['producto', 'nombre', 'nombre_producto'],
			'categoria': ['categoria', 'id_area'],
			'area': ['area'],
			'subarea': ['subarea', 'sub_area'],
			'unidad': ['unidad', 'unidad_medida'],
			'punto_minimo': ['punto_minimo', 'punto_min', 'minimo'],
			'stock_central': ['stock_central', 'stock_actual', 'stock'],
			'estado': ['estado'],
			'sede': ['sede', 'sede_nombre', 'nombre_sede'],
		}

		def idx(key):
			for alias in aliases[key]:
				if alias in header_map:
					return header_map[alias]
			return None

		def cell_value(row_values, column_idx):
			if column_idx is None or column_idx >= len(row_values):
				return None
			return row_values[column_idx]

		def ensure_categoria(nombre_categoria):
			nombre_categoria = (nombre_categoria or '').strip()
			if not nombre_categoria:
				return ''
			categoria = Categoria.query.filter(db.func.lower(Categoria.nombre_categoria) == nombre_categoria.lower()).first()
			if categoria:
				return categoria.nombre_categoria
			db.session.add(Categoria(nombre_categoria=nombre_categoria))
			return nombre_categoria

		def ensure_unidad(nombre_unidad):
			nombre_unidad = (nombre_unidad or '').strip()
			if not nombre_unidad:
				return ''
			unidad = Unidad.query.filter(db.func.lower(Unidad.nombre_unidad) == nombre_unidad.lower()).first()
			if unidad:
				return unidad.nombre_unidad
			db.session.add(Unidad(nombre_unidad=nombre_unidad))
			return nombre_unidad

		def ensure_area(nombre_area):
			nombre_area = _slugify(nombre_area)
			if not nombre_area:
				return ''
			area = Area.query.filter(db.func.lower(Area.nombre_area) == nombre_area).first()
			if area:
				return area.nombre_area
			db.session.add(Area(nombre_area=nombre_area))
			db.session.flush()
			return nombre_area

		def ensure_subarea(nombre_area, nombre_subarea):
			nombre_subarea = _slugify(nombre_subarea)
			if not nombre_subarea or not nombre_area:
				return ''
			area = Area.query.filter(db.func.lower(Area.nombre_area) == _slugify(nombre_area)).first()
			if not area:
				return ''
			exists = Subarea.query.filter_by(id_area=area.id_area, nombre_subarea=nombre_subarea).first()
			if not exists:
				db.session.add(Subarea(id_area=area.id_area, nombre_subarea=nombre_subarea))
			return nombre_subarea

		def ensure_sede(nombre_sede):
			nombre_sede = (nombre_sede or '').strip()
			if not nombre_sede:
				return None
			sede = Sede.query.filter(db.func.lower(Sede.nombre_sede) == nombre_sede.lower()).first()
			if sede:
				return sede
			sede = Sede(nombre_sede=nombre_sede)
			db.session.add(sede)
			db.session.flush()
			return sede

		id_idx = idx('id')
		name_idx = idx('producto')
		if id_idx is None or name_idx is None:
			flash('El Excel debe incluir columnas ID y Producto.', 'error')
			return redirect(url_for('inventario'))

		cat_idx = idx('categoria')
		area_idx = idx('area')
		sub_idx = idx('subarea')
		unit_idx = idx('unidad')
		min_idx = idx('punto_minimo')
		stock_idx = idx('stock_central')
		estado_idx = idx('estado')
		sede_idx = idx('sede')

		processed = 0
		deleted = 0
		errors = []
		imported_pairs = set()
		sedes_objetivo = set()

		try:
			for row_number, row in enumerate(rows, start=2):
				id_producto = str(cell_value(row, id_idx)).strip() if cell_value(row, id_idx) is not None else ''
				nombre_producto = str(cell_value(row, name_idx)).strip() if cell_value(row, name_idx) is not None else ''

				# Ignora filas completamente vacias
				if not id_producto and not nombre_producto:
					continue

				if not id_producto or not nombre_producto:
					errors.append(f'Fila {row_number}: ID y Producto son obligatorios.')
					continue

				sede_nombre = ''
				if sede_idx is not None and cell_value(row, sede_idx) is not None:
					sede_nombre = str(cell_value(row, sede_idx)).strip()

				if current_user.rol_nombre == 'admin_general':
					target_sede_obj = ensure_sede(sede_nombre) if sede_nombre else current_user.sede
					if not target_sede_obj:
						errors.append(f'Fila {row_number}: no se pudo resolver la sede.')
						continue
				else:
					target_sede_obj = current_user.sede

				sedes_objetivo.add(target_sede_obj.id_sede)

				categoria_val = str(cell_value(row, cat_idx)).strip() if cat_idx is not None and cell_value(row, cat_idx) is not None else ''
				if categoria_val:
					categoria_val = ensure_categoria(categoria_val)

				area_val_raw = str(cell_value(row, area_idx)).strip() if area_idx is not None and cell_value(row, area_idx) is not None else ''
				area_val = ensure_area(area_val_raw or 'cocina')

				subarea_val_raw = str(cell_value(row, sub_idx)).strip() if sub_idx is not None and cell_value(row, sub_idx) is not None else ''
				subarea_val = ensure_subarea(area_val, subarea_val_raw) if subarea_val_raw else ''

				unidad_val = str(cell_value(row, unit_idx)).strip() if unit_idx is not None and cell_value(row, unit_idx) is not None else ''
				if unidad_val:
					unidad_val = ensure_unidad(unidad_val)

				estado_val = str(cell_value(row, estado_idx)).strip() if estado_idx is not None and cell_value(row, estado_idx) is not None else 'Activo'
				estado_val = estado_val or 'Activo'

				producto = Producto.query.filter_by(id_producto=id_producto).first()
				if not producto:
					producto = Producto(id_producto=id_producto)
					db.session.add(producto)

				producto.nombre_producto = nombre_producto
				producto.id_area = categoria_val or producto.id_area or ''
				producto.area = area_val or 'cocina'
				if subarea_val:
					producto.subarea = subarea_val
				else:
					producto.subarea = _normalize_subarea(producto.area, producto.subarea)
				producto.unidad = unidad_val or producto.unidad or 'unidad'
				producto.estado = estado_val

				inv = InventarioSede.query.filter_by(id_sede=target_sede_obj.id_sede, id_producto=id_producto).first()
				if not inv:
					inv = InventarioSede(id_sede=target_sede_obj.id_sede, id_producto=id_producto)
					db.session.add(inv)

				if min_idx is not None:
					inv.punto_minimo = _safe_float(cell_value(row, min_idx), inv.punto_minimo or 0.0)
				if stock_idx is not None:
					inv.stock_actual = _safe_float(cell_value(row, stock_idx), inv.stock_actual or 0.0)

				imported_pairs.add((target_sede_obj.id_sede, id_producto))
				processed += 1

			if errors:
				db.session.rollback()
				preview = '; '.join(errors[:5])
				remaining = len(errors) - 5
				if remaining > 0:
					preview += f'; y {remaining} error(es) mas'
				flash(f'No se subio el Excel. Motivos: {preview}', 'error')
				return redirect(url_for('inventario'))

			if not sedes_objetivo:
				sedes_objetivo.add(current_user.id_sede)

			db.session.commit()
		except Exception as exc:
			db.session.rollback()
			flash(f'No se pudo subir el Excel: {exc}', 'error')
			return redirect(url_for('inventario'))

		flash(f'Importacion OK. Filas sincronizadas: {processed}. Registros eliminados por sincronizacion: {deleted}.', 'ok')
		return redirect(url_for('inventario'))

	@app.route('/movimientos', methods=['GET', 'POST'])
	@login_required
	def movimientos():
		if not current_user.can_view('movimientos'):
			return _forbidden_redirect()

		if request.method == 'POST':
			if not current_user.can_write('movimientos', 'insert'):
				return _forbidden_redirect()

			motivo = request.form.get('motivo', '').strip()
			motivo_nuevo = request.form.get('motivo_nuevo', '').strip()
			if motivo == 'OTRO':
				motivo = motivo_nuevo
			elif motivo == '':
				motivo = motivo_nuevo

			if not motivo:
				flash('Debes seleccionar o escribir un motivo.', 'error')
				return redirect(url_for('movimientos'))

			id_producto = request.form.get('id_producto', '').strip()
			cantidad = _safe_float(request.form.get('cantidad'), 0.0)
			tipo_movimiento = (request.form.get('tipo', 'ENTRADA') or 'ENTRADA').strip().upper()

			if not id_producto:
				flash('Debes seleccionar un producto valido.', 'error')
				return redirect(url_for('movimientos'))
			if cantidad <= 0:
				flash('La cantidad debe ser mayor a 0.', 'error')
				return redirect(url_for('movimientos'))
			if tipo_movimiento not in {'ENTRADA', 'SALIDA'}:
				flash('Tipo de movimiento invalido.', 'error')
				return redirect(url_for('movimientos'))

			producto = Producto.query.filter_by(id_producto=id_producto).first()
			if not producto:
				flash('El producto seleccionado no existe.', 'error')
				return redirect(url_for('movimientos'))

			inventario_row = InventarioSede.query.filter_by(
				id_sede=current_user.id_sede,
				id_producto=id_producto,
			).first()
			if not inventario_row:
				inventario_row = InventarioSede(
					id_sede=current_user.id_sede,
					id_producto=id_producto,
					stock_actual=0.0,
					punto_minimo=0.0,
				)
				db.session.add(inventario_row)

			stock_actual = _safe_float(inventario_row.stock_actual, 0.0)
			if tipo_movimiento == 'SALIDA' and cantidad > stock_actual:
				flash(f'Stock insuficiente. Disponible: {stock_actual:.2f}', 'error')
				return redirect(url_for('movimientos'))

			if tipo_movimiento == 'ENTRADA':
				inventario_row.stock_actual = stock_actual + cantidad
			else:
				inventario_row.stock_actual = stock_actual - cantidad

			db.session.add(
				MovimientoInventario(
					id_sede=current_user.id_sede,
					id_producto=id_producto,
					cantidad=cantidad,
					tipo=tipo_movimiento,
					motivo=motivo,
					fecha=datetime.utcnow(),
					id_usuario=current_user.id_usuario,
				)
			)
			db.session.commit()
			flash('Movimiento registrado y stock actualizado.', 'ok')

		q = request.args.get('q', '').strip()
		fecha_desde = request.args.get('fecha_desde', '').strip()
		fecha_hasta = request.args.get('fecha_hasta', '').strip()
		tipo = request.args.get('tipo', '').strip()
		categoria = request.args.get('categoria', '').strip()
		usuario_id = request.args.get('usuario_id', '').strip()

		movs_query = (
			db.session.query(MovimientoInventario, Usuario, Producto)
			.outerjoin(Usuario, Usuario.id_usuario == MovimientoInventario.id_usuario)
			.outerjoin(Producto, Producto.id_producto == MovimientoInventario.id_producto)
		)

		if current_user.rol_nombre != 'admin_general':
			movs_query = movs_query.filter(MovimientoInventario.id_sede == current_user.id_sede)

		if fecha_desde:
			movs_query = movs_query.filter(db.func.date(MovimientoInventario.fecha) >= fecha_desde)
		if fecha_hasta:
			movs_query = movs_query.filter(db.func.date(MovimientoInventario.fecha) <= fecha_hasta)
		if tipo:
			movs_query = movs_query.filter(MovimientoInventario.tipo == tipo)
		if categoria:
			movs_query = movs_query.filter(Producto.id_area == categoria)
		if usuario_id:
			movs_query = movs_query.filter(MovimientoInventario.id_usuario == usuario_id)
		if q:
			like_q = f"%{q}%"
			movs_query = movs_query.filter(
				or_(
					MovimientoInventario.id_producto.ilike(like_q),
					Producto.nombre_producto.ilike(like_q),
					Producto.id_area.ilike(like_q),
					MovimientoInventario.motivo.ilike(like_q),
					MovimientoInventario.tipo.ilike(like_q),
					Usuario.username.ilike(like_q),
				)
			)

		movs = movs_query.order_by(MovimientoInventario.id_movimiento.desc()).limit(300).all()

		usuarios_query = db.session.query(Usuario.id_usuario, Usuario.username).join(
			MovimientoInventario, MovimientoInventario.id_usuario == Usuario.id_usuario
		).distinct()
		if current_user.rol_nombre != 'admin_general':
			usuarios_query = usuarios_query.filter(MovimientoInventario.id_sede == current_user.id_sede)
		usuarios_filtro = usuarios_query.order_by(Usuario.username.asc()).all()
		categorias_filtro = [
			categoria.nombre_categoria
			for categoria in Categoria.query.order_by(Categoria.nombre_categoria.asc()).all()
		]
		categorias_extra = [
			row[0]
			for row in db.session.query(Producto.id_area)
				.filter(Producto.id_area.isnot(None), Producto.id_area != '')
				.distinct()
				.order_by(Producto.id_area.asc())
				.all()
			if row[0] not in categorias_filtro
		]
		categorias_filtro.extend(categorias_extra)
		productos_query = db.session.query(Producto).join(
			InventarioSede,
			InventarioSede.id_producto == Producto.id_producto,
		)
		if current_user.rol_nombre != 'admin_general':
			productos_query = productos_query.filter(InventarioSede.id_sede == current_user.id_sede)
		productos = productos_query.distinct().order_by(Producto.nombre_producto.asc()).all()
		stock_rows = InventarioSede.query.filter_by(id_sede=current_user.id_sede).all()
		stock_por_producto = {row.id_producto: _safe_float(row.stock_actual, 0.0) for row in stock_rows}

		return render_template(
			'dashboard/movimientos.html',
			allowed_views=_allowed_views(current_user),
			movimientos=movs,
			productos=productos,
			stock_por_producto=stock_por_producto,
			categorias_filtro=categorias_filtro,
			usuarios_filtro=usuarios_filtro,
			selected_q=q,
			selected_fecha_desde=fecha_desde,
			selected_fecha_hasta=fecha_hasta,
			selected_tipo=tipo,
			selected_categoria=categoria,
			selected_usuario_id=usuario_id,
			can_insert=current_user.can_write('movimientos', 'insert'),
		)

	@app.route('/pedidos', methods=['GET', 'POST'])
	@login_required
	def pedidos():
		if not current_user.can_view('pedidos'):
			return _forbidden_redirect()
		selected_date = _get_selected_app_date()
		can_update = current_user.can_write('pedidos', 'update')
		can_delete_requested = current_user.rol_nombre in {'admin_general', 'admin_almacen'}

		pedido_id_raw = request.args.get('pedido_id', '').strip()
		pedido_id = int(pedido_id_raw) if pedido_id_raw.isdigit() else None
		is_async_request = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

		def _pedidos_post_response(target_pedido_id=None):
			if is_async_request:
				return ('', 204)
			return redirect(
				url_for(
					'pedidos',
					pedido_id=target_pedido_id if target_pedido_id else pedido_id,
					scroll_y=form_scroll if form_scroll else None,
				)
			)

		if request.method == 'POST':
			action = request.form.get('action', '').strip()
			form_scroll = request.form.get('scroll_y', '').strip()
			if action == 'mark_sent':
				if not can_update:
					return _forbidden_redirect()
				pedido = ChecklistPedido.query.get(request.form.get('id_pedido'))
				if pedido and pedido.estado_general in {'Pendiente', 'Borrador'}:
					total = DetallePedido.query.filter(
						DetallePedido.id_pedido == pedido.id_pedido,
						DetallePedido.cantidad_pedida > 0,
					).count()
					if total == 0:
						flash('El pedido no tiene lineas para enviar.', 'error')
					else:
						pedido.estado_general = 'Enviado'
						db.session.commit()
						flash('Pedido enviado a sede. Cocina confirmara solo las lineas enviadas.', 'ok')
				else:
					flash('No se pudo actualizar el pedido.', 'error')
				return _pedidos_post_response(pedido.id_pedido if pedido else pedido_id)
			elif action == 'save_dispatch_line':
				if not can_update:
					return _forbidden_redirect()
				detalle = DetallePedido.query.get(request.form.get('id_detalle'))
				if not detalle:
					flash('Linea no encontrada.', 'error')
					return _pedidos_post_response(request.form.get('pedido_id'))

				pedido = ChecklistPedido.query.get(detalle.id_pedido)
				cantidad = _safe_float(request.form.get('cantidad_entregada'), detalle.cantidad_pedida or 0.0)
				cantidad = max(cantidad, 0.0)
				checked = request.form.get('enviar_linea') == 'on'
				if checked and cantidad <= 0:
					cantidad = max(detalle.cantidad_pedida or 0.0, 1.0)

				cantidad_anterior = _safe_float(detalle.cantidad_entregada, 0.0)
				cantidad_nueva = cantidad if checked else 0.0
				delta = cantidad_nueva - cantidad_anterior

				try:
					_apply_dispatch_inventory_delta(
						current_user.id_sede,
						detalle.id_producto,
						delta,
						detalle.id_pedido,
						detalle.id_detalle,
					)
				except ValueError as exc:
					db.session.rollback()
					flash(str(exc), 'error')
					return _pedidos_post_response(request.form.get('pedido_id'))

				detalle.cantidad_entregada = cantidad_nueva
				if detalle.estado_sede != 'Recibido':
					detalle.estado_sede = 'Pendiente'
				db.session.commit()
				return _pedidos_post_response(request.form.get('pedido_id'))
			elif action == 'delete_requested_order':
				if not can_delete_requested:
					return _forbidden_redirect()
				pedido = ChecklistPedido.query.get(request.form.get('id_pedido'))
				if not pedido:
					flash('Pedido no encontrado.', 'error')
					return _pedidos_post_response(pedido_id)

				deleted_sede_id = pedido.id_sede
				deleted_turno_id = pedido.id_turno
				DetallePedido.query.filter_by(id_pedido=pedido.id_pedido).delete(synchronize_session=False)
				db.session.delete(pedido)

				# Si ya no quedan pedidos cerrados para ese alcance y fecha,
				# vuelve a crear lista editable desde las plantillas de cada usuario.
				scope_users = Usuario.query.join(Rol, Rol.id_rol == Usuario.id_rol).filter(
					Usuario.id_sede == deleted_sede_id,
					Usuario.id_turno == deleted_turno_id,
					Rol.nombre_rol.in_(['cocinero', 'admin_sala']),
				).all()
				for scope_user in scope_users:
					_build_checklist_from_template_if_needed(scope_user, selected_date)

				db.session.commit()
				flash('Pedido eliminado correctamente. Cocina puede volver a generar su lista.', 'ok')
				return _pedidos_post_response(None)
			else:
				if not current_user.can_write('pedidos', 'insert'):
					return _forbidden_redirect()

				pedido = ChecklistPedido(
					id_sede=current_user.id_sede,
					id_turno=current_user.id_turno,
					id_usuario=current_user.id_usuario,
					estado_general='Pendiente',
				)
				db.session.add(pedido)
				db.session.flush()

				db.session.add(
					DetallePedido(
						id_pedido=pedido.id_pedido,
						id_usuario=current_user.id_usuario,
						id_producto=request.form.get('id_producto'),
						cantidad_pedida=float(request.form.get('cantidad_pedida', 0)),
					)
				)
				db.session.commit()
				flash('Pedido creado.', 'ok')
				if is_async_request:
					return ('', 204)
				return redirect(url_for('pedidos', pedido_id=pedido.id_pedido))

		pedidos_query = db.session.query(ChecklistPedido, Sede, Turno).outerjoin(
			Sede, Sede.id_sede == ChecklistPedido.id_sede
		).outerjoin(
			Turno, Turno.id_turno == ChecklistPedido.id_turno
		)
		pedidos_query = pedidos_query.filter(db.func.date(ChecklistPedido.fecha) == selected_date)
		if current_user.rol_nombre not in {'admin_general', 'admin_almacen', 'personal_prod'}:
			pedidos_query = pedidos_query.filter(ChecklistPedido.id_sede == current_user.id_sede)
		pedidos_query = pedidos_query.order_by(ChecklistPedido.id_pedido.desc())
		pedido_rows = pedidos_query.limit(80).all()

		if pedido_id is None and pedido_rows:
			pedido_id = pedido_rows[0][0].id_pedido

		selected_pedido = None
		selected_items = []
		selected_sede_nombre = ''
		selected_turno_nombre = ''
		if pedido_id is not None:
			selected_pedido = ChecklistPedido.query.get(pedido_id)
			if selected_pedido:
				selected_items_query = db.session.query(DetallePedido, Producto).join(
					Producto, Producto.id_producto == DetallePedido.id_producto
				).filter(
					DetallePedido.id_pedido == selected_pedido.id_pedido,
					DetallePedido.cantidad_pedida > 0,
				)
				if current_user.rol_nombre not in {'admin_general', 'admin_almacen', 'personal_prod'}:
					selected_items_query = selected_items_query.filter(
						or_(
							DetallePedido.id_usuario == current_user.id_usuario,
							DetallePedido.id_usuario.is_(None),
						)
					)
				selected_items = selected_items_query.order_by(Producto.nombre_producto.asc()).all()

				# Nombre legible de sede/turno para la UI de impresión
				if selected_pedido:
					if selected_pedido.id_sede:
						sede_obj = Sede.query.get(selected_pedido.id_sede)
						selected_sede_nombre = sede_obj.nombre_sede if sede_obj else ''
					if selected_pedido.id_turno:
						turno_obj = Turno.query.get(selected_pedido.id_turno)
						selected_turno_nombre = turno_obj.nombre_turno if turno_obj else ''

		pedido_area_map = {}
		for pedido_obj, _, _ in pedido_rows:
			areas = db.session.query(Producto.area).join(
				DetallePedido, DetallePedido.id_producto == Producto.id_producto
			).filter(
				DetallePedido.id_pedido == pedido_obj.id_pedido
			).distinct().all()
			area_labels = [(_normalize_area(area_name) or area_name or '').title() for (area_name,) in areas if area_name]
			pedido_area_map[pedido_obj.id_pedido] = ', '.join(area_labels) if area_labels else '-'

		return render_template(
			'dashboard/pedidos.html',
			allowed_views=_allowed_views(current_user),
			pedido_rows=pedido_rows,
			selected_pedido=selected_pedido,
			selected_items=selected_items,
			pedido_area_map=pedido_area_map,
			productos=Producto.query.order_by(Producto.nombre_producto).all(),
				can_insert=current_user.can_write('pedidos', 'insert'),
				selected_sede_nombre=selected_sede_nombre,
				selected_turno_nombre=selected_turno_nombre,
			can_update=can_update,
			can_delete_requested=can_delete_requested,
		)

	@app.route('/pedidos/<int:id_pedido>/imprimir', methods=['GET'])
	@login_required
	def imprimir_pedido(id_pedido):
		if not current_user.can_view('pedidos'):
			return _forbidden_redirect()

		pedido = db.session.get(ChecklistPedido, id_pedido)
		if not pedido:
			flash('Pedido no encontrado.', 'error')
			return redirect(url_for('pedidos'))
		if current_user.rol_nombre not in {'admin_general', 'admin_almacen', 'personal_prod'} and pedido.id_sede != current_user.id_sede:
			return _forbidden_redirect()

		formato = request.args.get('formato', 'A4')
		if formato not in {'A4', '80mm'}:
			formato = 'A4'
		items = db.session.query(DetallePedido, Producto).join(
			Producto, Producto.id_producto == DetallePedido.id_producto
		).filter(
			DetallePedido.id_pedido == pedido.id_pedido,
			DetallePedido.cantidad_pedida > 0,
		).order_by(Producto.nombre_producto.asc()).all()
		sede = db.session.get(Sede, pedido.id_sede) if pedido.id_sede else None
		turno = db.session.get(Turno, pedido.id_turno) if pedido.id_turno else None
		return render_template(
			'dashboard/pedido_print.html',
			pedido=pedido,
			items=items,
			sede=sede,
			turno=turno,
			formato=formato,
		)

	@app.route('/checklist', methods=['GET', 'POST'])
	@login_required
	def checklist():
		if not current_user.can_view('checklist'):
			return _forbidden_redirect()
		selected_date = _get_selected_app_date()
		is_admin_general = current_user.rol_nombre == 'admin_general'
		selected_pedido_raw = request.args.get('pedido_id', '').strip()
		selected_pedido_id = int(selected_pedido_raw) if selected_pedido_raw.isdigit() else None
		checklist_selector_options = []
		selected_filter_turno = request.args.get('f_turno', '').strip()
		selected_filter_sede = request.args.get('f_sede', '').strip()
		selected_filter_area = request.args.get('f_area', '').strip().lower()
		selected_filter_user = request.args.get('f_user', '').strip()
		is_async_request = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
		admin_turno_options = []
		admin_sede_options = []
		admin_area_options = []
		admin_user_options = []
		checklist_user_options = []
		can_edit_selected_user = True

		if not is_admin_general:
			seeded_checklist = _build_checklist_from_template_if_needed(current_user, selected_date)
			if seeded_checklist is not None:
				db.session.commit()
			scope_users = _checklist_scope_users(current_user)
			checklist_user_options = [
				{'id': scope_user.id_usuario, 'label': scope_user.username}
				for scope_user in sorted(scope_users, key=lambda item: (item.username or '').lower())
			]
			allowed_user_ids = {option['id'] for option in checklist_user_options}
			if not selected_filter_user or selected_filter_user not in allowed_user_ids:
				selected_filter_user = current_user.id_usuario

		active_checklist = None
		visible_checklist = None
		if is_admin_general:
			admin_rows = db.session.query(ChecklistPedido, Sede, Turno, Usuario).outerjoin(
				Sede, Sede.id_sede == ChecklistPedido.id_sede
			).outerjoin(
				Turno, Turno.id_turno == ChecklistPedido.id_turno
			).outerjoin(
				Usuario, Usuario.id_usuario == ChecklistPedido.id_usuario
			).filter(
				db.func.date(ChecklistPedido.fecha) == selected_date.strftime('%Y-%m-%d')
			).order_by(ChecklistPedido.id_pedido.desc()).limit(200).all()

			turno_map = {}
			for _, _, turno, _ in admin_rows:
				if turno:
					turno_map[turno.id_turno] = turno.nombre_turno
			admin_turno_options = [
				{'id': turno_id, 'label': f"{turno_label} ({turno_id})"}
				for turno_id, turno_label in sorted(turno_map.items(), key=lambda item: (item[1] or '', item[0] or ''))
			]
			if not selected_filter_turno and admin_turno_options:
				selected_filter_turno = admin_turno_options[0]['id']

			rows_by_turno = [row for row in admin_rows if not selected_filter_turno or (row[2] and row[2].id_turno == selected_filter_turno)]

			sede_map = {}
			for _, sede, _, _ in rows_by_turno:
				if sede:
					sede_map[sede.id_sede] = sede.nombre_sede
			admin_sede_options = [
				{'id': str(sede_id), 'label': sede_name}
				for sede_id, sede_name in sorted(sede_map.items(), key=lambda item: (item[1] or '', item[0]))
			]
			if not selected_filter_sede and admin_sede_options:
				selected_filter_sede = admin_sede_options[0]['id']

			rows_by_sede = [
				row for row in rows_by_turno
				if not selected_filter_sede or (row[1] and str(row[1].id_sede) == selected_filter_sede)
			]
			pedido_ids_by_sede = [pedido.id_pedido for pedido, _, _, _ in rows_by_sede]

			for pedido, sede, turno, usuario in rows_by_sede:
				checklist_selector_options.append({
					'id_pedido': pedido.id_pedido,
					'label': f"#{pedido.id_pedido} | {usuario.username if usuario else '-'} | {sede.nombre_sede if sede else '-'} | {turno.nombre_turno if turno else '-'} | {pedido.estado_general}",
				})

			if selected_pedido_id and selected_pedido_id in pedido_ids_by_sede:
				visible_checklist = ChecklistPedido.query.get(selected_pedido_id)
			elif rows_by_sede:
				visible_checklist = rows_by_sede[0][0]
				selected_pedido_id = visible_checklist.id_pedido
			else:
				selected_pedido_id = None

			if pedido_ids_by_sede:
				area_rows = db.session.query(Producto.area).join(
					DetallePedido, DetallePedido.id_producto == Producto.id_producto
				).filter(
					DetallePedido.id_pedido.in_(pedido_ids_by_sede)
				).distinct().all()
				area_values = [(_slugify(area_name) if area_name else '') for (area_name,) in area_rows if area_name]
				admin_area_options = [
					{'id': area_value, 'label': area_value.replace('_', ' ').title()}
					for area_value in sorted(set(area_values))
				]
				if selected_filter_area and selected_filter_area not in {option['id'] for option in admin_area_options}:
					selected_filter_area = ''
				if not selected_filter_area and admin_area_options:
					selected_filter_area = admin_area_options[0]['id']

				user_query = db.session.query(Usuario.id_usuario, Usuario.username).join(
					DetallePedido, DetallePedido.id_usuario == Usuario.id_usuario
				).join(
					Producto, Producto.id_producto == DetallePedido.id_producto
				).filter(
					DetallePedido.id_pedido.in_(pedido_ids_by_sede)
				)
				if selected_filter_area:
					user_query = user_query.filter(db.func.lower(Producto.area) == selected_filter_area)
				user_rows = user_query.distinct().all()
				admin_user_options = [
					{'id': user_id, 'label': username}
					for user_id, username in sorted(user_rows, key=lambda item: (item[1] or '').lower())
				]
				if selected_filter_user and selected_filter_user not in {option['id'] for option in admin_user_options}:
					selected_filter_user = ''
				if not selected_filter_user and admin_user_options:
					selected_filter_user = admin_user_options[0]['id']

			if visible_checklist and visible_checklist.estado_general in {'Borrador', 'Pendiente', 'Enviado'}:
				active_checklist = visible_checklist
		else:
			active_checklist = _get_active_checklist(current_user, selected_date)
			visible_checklist = active_checklist or _get_visible_checklist(current_user, selected_date)

		if request.method == 'POST':
			action = request.form.get('action', '').strip()
			if not current_user.can_write('checklist', 'insert'):
				return _forbidden_redirect()

			is_viewing_other_user = bool(
				not is_admin_general and selected_filter_user and selected_filter_user != current_user.id_usuario
			)
			if is_viewing_other_user and action in {
				'add_item', 'remove_selected', 'qty_plus', 'qty_minus', 'qty_clear', 'qty_set', 'remove_item', 'send_list', 'confirm_item'
			}:
				flash('Solo puedes editar tu propia lista. La vista de otro usuario es solo lectura.', 'error')
				db.session.rollback()
				return redirect(
					url_for(
						'checklist',
						tab=request.form.get('next_tab', request.args.get('tab', 'view')).strip() or 'view',
						q=request.form.get('q', request.args.get('q', '')).strip(),
						f_user=selected_filter_user or None,
					)
				)

			if action == 'import_template':
				if not can_edit_selected_user:
					return _forbidden_redirect()

				upload = request.files.get('template_file')
				if not upload or not upload.filename:
					flash('Debes seleccionar un archivo JSON.', 'error')
					return redirect(url_for('checklist', tab='edit', f_user=selected_filter_user or None))

				try:
					payload = json.loads(upload.read().decode('utf-8'))
				except Exception:
					flash('El archivo no es un JSON valido.', 'error')
					return redirect(url_for('checklist', tab='edit', f_user=selected_filter_user or None))

				target_user = current_user
				if is_admin_general and selected_filter_user:
					target_user = db.session.get(Usuario, selected_filter_user) or current_user

				result = _replace_template_from_import(target_user, target_user.id_usuario, payload)
				db.session.commit()
				_build_checklist_from_template_if_needed(target_user, selected_date)
				_sync_open_checklists_with_template(target_user, selected_date)
				flash(
					f"Plantilla importada. Productos agregados: {result['importados']}. Omitidos: {result['omitidos']}",
					'ok',
				)
				return redirect(url_for('checklist', tab='edit', f_user=selected_filter_user or None))

			def _get_target_detail():
				detail_query = DetallePedido.query.filter_by(
					id_detalle=request.form.get('id_detalle', '').strip(),
					id_pedido=active_checklist.id_pedido if active_checklist else None,
				)
				if is_admin_general and selected_filter_user:
					detail_query = detail_query.filter_by(id_usuario=selected_filter_user)
				else:
					detail_query = detail_query.filter_by(id_usuario=current_user.id_usuario)
				return detail_query.first()

			if action == 'add_item':
				id_producto = request.form.get('id_producto', '').strip()
				if not id_producto:
					flash('Producto invalido.', 'error')
				elif is_admin_general and active_checklist and selected_filter_user:
					exists = DetallePedido.query.filter_by(
						id_pedido=active_checklist.id_pedido,
						id_usuario=selected_filter_user,
						id_producto=id_producto,
					).first()
					if exists:
						flash('Ese producto ya esta en la lista de ese usuario.', 'ok')
					else:
						db.session.add(
							DetallePedido(
								id_pedido=active_checklist.id_pedido,
								id_usuario=selected_filter_user,
								id_producto=id_producto,
								cantidad_pedida=0.0,
								estado_sede='Pendiente',
							)
						)
						flash('Producto agregado a la lista del usuario seleccionado.', 'ok')
				else:
					exists = _template_scope_query(current_user).filter_by(id_producto=id_producto).first()
					if exists:
						flash('Ese producto ya esta en tu plantilla.', 'ok')
					else:
						db.session.add(
							PlantillaChecklistItem(
								id_usuario=current_user.id_usuario,
								id_sede=current_user.id_sede,
								id_turno=current_user.id_turno,
								area=_preferred_area_for_user(current_user),
								id_producto=id_producto,
							)
						)
						flash('Producto agregado a tu plantilla personal.', 'ok')
				if not is_admin_general:
					_build_checklist_from_template_if_needed(current_user, selected_date)
					_sync_open_checklists_with_template(current_user, selected_date)

			elif action == 'remove_selected':
				id_producto = request.form.get('id_producto', '').strip()
				if not id_producto:
					flash('Producto invalido.', 'error')
				elif is_admin_general and active_checklist and selected_filter_user:
					item = DetallePedido.query.filter_by(
						id_pedido=active_checklist.id_pedido,
						id_usuario=selected_filter_user,
						id_producto=id_producto,
					).first()
					if item:
						db.session.delete(item)
						flash('Producto quitado de la lista del usuario seleccionado.', 'ok')
				else:
					removed_count = _template_scope_query(current_user).filter_by(id_producto=id_producto).delete(synchronize_session=False)
					if removed_count:
						flash('Producto quitado de tu plantilla personal.', 'ok')
					if not is_admin_general:
						_build_checklist_from_template_if_needed(current_user, selected_date)
						_sync_open_checklists_with_template(current_user, selected_date)

			elif action == 'qty_plus':
				if not active_checklist or active_checklist.estado_general not in {'Borrador', 'Pendiente'}:
					flash('La lista ya fue enviada.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle:
						detalle.cantidad_pedida = max((detalle.cantidad_pedida or 0.0) + 1.0, 0.0)

			elif action == 'qty_minus':
				if not active_checklist or active_checklist.estado_general not in {'Borrador', 'Pendiente'}:
					flash('La lista ya fue enviada.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle:
						actual = max(_safe_float(detalle.cantidad_pedida, 0.0), 0.0)
						if actual > 1.0:
							detalle.cantidad_pedida = actual - 1.0

			elif action == 'qty_clear':
				if not active_checklist or active_checklist.estado_general not in {'Borrador', 'Pendiente'}:
					flash('La lista ya fue enviada.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle:
						detalle.cantidad_pedida = 0.0

			elif action == 'qty_set':
				if not active_checklist or active_checklist.estado_general not in {'Borrador', 'Pendiente'}:
					flash('La lista ya fue enviada.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle:
						cantidad = max(_safe_float(request.form.get('cantidad_pedida'), detalle.cantidad_pedida or 0.0), 0.0)
						detalle.cantidad_pedida = cantidad

			elif action == 'remove_item':
				if not active_checklist or active_checklist.estado_general not in {'Borrador', 'Pendiente'}:
					flash('La lista ya fue enviada.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle:
						db.session.delete(detalle)
						flash('Producto quitado de la lista.', 'ok')

			elif action == 'send_list':
				if not active_checklist:
					flash('No hay lista para enviar.', 'error')
				elif active_checklist.estado_general != 'Borrador':
					flash('La lista ya fue enviada.', 'error')
				elif DetallePedido.query.filter(
					DetallePedido.id_pedido == active_checklist.id_pedido,
					DetallePedido.cantidad_pedida > 0,
				).count() == 0:
					flash('Agrega productos antes de enviar.', 'error')
				else:
					active_checklist.estado_general = 'Pendiente'
					flash('Lista enviada. Esperando a almacén.', 'ok')

			elif action == 'confirm_item':
				if not active_checklist or active_checklist.estado_general != 'Enviado':
					flash('Aun no puedes confirmar recepción.', 'error')
				else:
					detalle = _get_target_detail()
					if detalle and (detalle.cantidad_entregada or 0) <= 0:
						flash('Ese item no fue enviado por almacén.', 'error')
					elif detalle and detalle.estado_sede != 'Recibido':
						_apply_dispatch_inventory_delta(
							active_checklist.id_sede,
							detalle.id_producto,
							-(_safe_float(detalle.cantidad_entregada, 0.0)),
							detalle.id_pedido,
							detalle.id_detalle,
						)
						detalle.estado_sede = 'Recibido'
						_complete_checklist_if_all_received(active_checklist)
						flash('Item recibido confirmado.', 'ok')

			db.session.commit()
			scroll_y = request.form.get('scroll_y', '').strip()
			if is_async_request:
				return ('', 204)
			return redirect(
				url_for(
					'checklist',
					tab=request.form.get('next_tab', request.args.get('tab', 'view')).strip() or 'view',
					q=request.form.get('q', request.args.get('q', '')).strip(),
					scroll_y=scroll_y if scroll_y else None,
					pedido_id=request.form.get('pedido_id', str(active_checklist.id_pedido if active_checklist else '')).strip() if is_admin_general else None,
					f_turno=request.form.get('f_turno', selected_filter_turno).strip() if is_admin_general else None,
					f_sede=request.form.get('f_sede', selected_filter_sede).strip() if is_admin_general else None,
					f_area=request.form.get('f_area', selected_filter_area).strip() if is_admin_general else None,
					f_user=request.form.get('f_user', selected_filter_user).strip() or None,
				)
			)

		if not is_admin_general:
			active_checklist = _get_active_checklist(current_user, selected_date)
			visible_checklist = active_checklist or _get_visible_checklist(current_user, selected_date)

		checklist_items = _get_checklist_items(
			visible_checklist,
			include_all=is_admin_general,
			user=current_user,
			target_user_id=selected_filter_user if (is_admin_general or selected_filter_user) else '',
			target_area=selected_filter_area if is_admin_general else '',
		)
		active_items = _get_checklist_items(
			active_checklist,
			include_all=is_admin_general,
			user=current_user,
			target_user_id=selected_filter_user if (is_admin_general or selected_filter_user) else '',
			target_area=selected_filter_area if is_admin_general else '',
		)
		active_positive_items = [row for row in active_items if _safe_float(row[0].cantidad_pedida, 0.0) > 0]
		visible_positive_items = [row for row in checklist_items if _safe_float(row[0].cantidad_pedida, 0.0) > 0]
		if is_admin_general and active_checklist and selected_filter_user:
			selected_product_ids = {
				row.id_producto
				for row in DetallePedido.query.filter_by(
					id_pedido=active_checklist.id_pedido,
					id_usuario=selected_filter_user,
				).all()
			}
		elif not is_admin_general and selected_filter_user and selected_filter_user != current_user.id_usuario:
			selected_product_ids = _get_template_product_ids_for_user(current_user, selected_filter_user)
		else:
			selected_product_ids = _get_template_product_ids(current_user)
		can_edit_selected_user = is_admin_general or not selected_filter_user or selected_filter_user == current_user.id_usuario
		selected_q = request.args.get('q', '').strip()
		catalog_sede_id = current_user.id_sede
		if is_admin_general and selected_filter_sede:
			try:
				catalog_sede_id = int(selected_filter_sede)
			except (TypeError, ValueError):
				catalog_sede_id = current_user.id_sede
		catalog_products = _get_checklist_catalog(current_user, selected_q, id_sede=catalog_sede_id)
		all_catalog_products = _get_checklist_catalog(current_user, '', id_sede=catalog_sede_id)
		active_tab = request.args.get('tab', 'view').strip().lower() or 'view'
		if active_tab not in {'view', 'list', 'edit'}:
			active_tab = 'view'
		return render_template(
			'dashboard/checklist.html',
			allowed_views=_allowed_views(current_user),
			current_checklist=visible_checklist,
			active_checklist=active_checklist,
			checklist_items=checklist_items,
			active_items=active_items,
			active_positive_items=active_positive_items,
			visible_positive_items=visible_positive_items,
			selected_product_ids=selected_product_ids,
			catalog_products=catalog_products,
			all_catalog_products=all_catalog_products,
			active_tab=active_tab,
			selected_q=selected_q,
			is_admin_general=is_admin_general,
			checklist_selector_options=checklist_selector_options,
			selected_pedido_id=selected_pedido_id,
			selected_filter_turno=selected_filter_turno,
			selected_filter_sede=selected_filter_sede,
			selected_filter_area=selected_filter_area,
			selected_filter_user=selected_filter_user,
			admin_turno_options=admin_turno_options,
			admin_sede_options=admin_sede_options,
			admin_area_options=admin_area_options,
			admin_user_options=admin_user_options,
			checklist_user_options=checklist_user_options,
			can_edit_selected_user=can_edit_selected_user,
			can_insert=current_user.can_write('checklist', 'insert'),
		)

	@app.route('/manifest.webmanifest')
	def manifest():
		response = send_from_directory(app.static_folder, 'manifest.webmanifest')
		response.headers['Content-Type'] = 'application/manifest+json'
		response.headers['Cache-Control'] = 'no-cache'
		return response

	@app.route('/service-worker.js')
	def service_worker():
		response = send_from_directory(app.static_folder, 'js/service-worker.js')
		response.headers['Content-Type'] = 'application/javascript'
		response.headers['Service-Worker-Allowed'] = '/'
		response.headers['Cache-Control'] = 'no-cache'
		return response

	@app.route('/horarios')
	@login_required
	def horarios():
		if not current_user.can_view('horarios'):
			return _forbidden_redirect()
		return render_template('horarios/agenda.html', allowed_views=_allowed_views(current_user), agenda_role=current_user.rol_nombre, agenda_sede_id=current_user.id_sede, agenda_turno_id=current_user.id_turno)

	@app.route('/horarios/<path:filename>')
	@login_required
	def horarios_asset(filename):
		if not current_user.can_view('horarios'):
			return _forbidden_redirect()
		return send_from_directory(os.path.join(app.template_folder, 'horarios'), filename)

	@app.route('/api/horarios/catalogos')
	@login_required
	def horarios_catalogos():
		if not current_user.can_view('horarios'):
			return jsonify({'error': 'forbidden'}), 403
		turnos = Turno.query.filter(Turno.id_turno != 'NA').order_by(Turno.nombre_turno).all()
		return jsonify({
			'sedes': [{'id': sede.id_sede, 'nombre': sede.nombre_sede, 'estado': 'activo'} for sede in Sede.query.order_by(Sede.nombre_sede).all()],
			'turnos': [{'id': index + 1, 'id_global': turno.id_turno, 'nombre': turno.nombre_turno, 'horaInicio': '12:00' if turno.id_turno == 'MANANA' else '18:30', 'horaFin': '17:30' if turno.id_turno == 'MANANA' else '00:30', 'toleranciaMinutos': 10, 'estado': 'activo'} for index, turno in enumerate(turnos)],
		})

	@app.route('/api/horarios/auditoria', methods=['POST'])
	@login_required
	def horarios_auditoria():
		if not current_user.can_write('horarios', 'insert'):
			return jsonify({'error': 'forbidden'}), 403
		payload = request.get_json(silent=True) or {}
		db.session.add(AgendaAuditoria(id_usuario=current_user.id_usuario, accion=str(payload.get('accion', 'cambio'))[:40], entidad=str(payload.get('entidad', 'agenda'))[:40], detalle_json=json.dumps(payload, ensure_ascii=True, default=str)))
		db.session.commit()
		return jsonify({'ok': True})

	@app.route('/api/horarios/datos', methods=['GET', 'PUT'])
	@login_required
	def horarios_datos():
		if not current_user.can_view('horarios'):
			return jsonify({'error': 'forbidden'}), 403
		registro = db.session.get(AgendaPersistencia, 1)
		if request.method == 'GET':
			return jsonify({'exists': bool(registro and registro.datos_json and registro.datos_json != '{}'), 'datos': json.loads(registro.datos_json) if registro else None})
		if not current_user.can_write('horarios', 'update'):
			return jsonify({'error': 'forbidden'}), 403
		payload = request.get_json(silent=True) or {}
		if not isinstance(payload, dict):
			return jsonify({'error': 'invalid_payload'}), 400
		if registro is None:
			registro = AgendaPersistencia(id_agenda=1)
			db.session.add(registro)
		registro.datos_json = json.dumps(payload, ensure_ascii=False, default=str)
		registro.actualizado_por = current_user.id_usuario
		db.session.add(AgendaAuditoria(id_usuario=current_user.id_usuario, accion='sincronizar_datos', entidad='agenda', detalle_json=json.dumps({'colecciones': list(payload.keys())}, ensure_ascii=True)))
		db.session.commit()
		return jsonify({'ok': True})

	@app.route('/api/horarios/exportar-trabajadores')
	@login_required
	def horarios_exportar_trabajadores():
		if not current_user.can_view('horarios'):
			return jsonify({'error': 'forbidden'}), 403
		registro = db.session.get(AgendaPersistencia, 1)
		datos = json.loads(registro.datos_json) if registro and registro.datos_json else {}
		trabajadores = datos.get('trabajadores', []) if isinstance(datos, dict) else []
		if current_user.rol_nombre != 'admin_general':
			turno_local = {'MANANA': 1, 'NOCHE': 2}.get(current_user.id_turno, current_user.id_turno)
			trabajadores = [item for item in trabajadores if str(item.get('sedeId')) == str(current_user.id_sede) and str(item.get('turnoId')) in {str(turno_local), str(current_user.id_turno)}]
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Trabajadores'
		columns = ['Nombre', 'Apellido', 'DNI', 'Telefono', 'Fecha nacimiento', 'Fecha ingreso', 'Direccion', 'Emergencia', 'Grado profesional', 'Profesion', 'Institucion de estudios', 'Area', 'Cargo principal', 'Otros cargos', 'Sede', 'Turno', 'Dia descanso', 'Estado']
		worksheet.append(columns)
		for trabajador in trabajadores:
			fila = [trabajador.get(key, '') for key in ('nombre', 'apellido', 'dni', 'telefono', 'fechaNacimiento', 'fechaIngreso', 'direccion', 'emergenciaNumero', 'gradoProfesional', 'profesion', 'institucionEstudios', 'areaId', 'cargos', 'otrosCargos', 'sedeId', 'turnoId', 'diaDescanso', 'estado')]
			fila[12] = ';'.join(str(value) for value in fila[12]) if isinstance(fila[12], list) else fila[12]
			fila[13] = ';'.join(str(value) for value in fila[13]) if isinstance(fila[13], list) else fila[13]
			worksheet.append(fila)
		buffer = BytesIO()
		workbook.save(buffer)
		buffer.seek(0)
		return send_file(buffer, as_attachment=True, download_name=f'trabajadores_{datetime.utcnow().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	@app.route('/api/horarios/molde-trabajadores')
	@login_required
	def horarios_molde_trabajadores():
		if not current_user.can_write('horarios', 'update'):
			return jsonify({'error': 'forbidden'}), 403
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = 'Trabajadores'
		columns = ['Nombre', 'Apellido', 'DNI', 'Telefono', 'Fecha nacimiento', 'Fecha ingreso', 'Direccion', 'Emergencia', 'Grado profesional', 'Profesion', 'Institucion de estudios', 'Area', 'Cargo principal', 'Otros cargos', 'Sede', 'Turno', 'Dia descanso', 'Estado']
		worksheet.append(columns)
		for cell in worksheet[1]:
			cell.font = Font(bold=True)
		worksheet.freeze_panes = 'A2'
		catalog_sheet = workbook.create_sheet('Catalogos')
		catalog_sheet.append(['Campo', 'Valores validos'])
		catalog_sheet.append(['Sede', ' | '.join(sede.nombre_sede for sede in Sede.query.order_by(Sede.nombre_sede).all())])
		catalog_sheet.append(['Turno', ' | '.join(turno.nombre_turno for turno in Turno.query.filter(Turno.id_turno != 'NA').order_by(Turno.nombre_turno).all())])
		catalog_sheet.append(['Area', 'Usa los nombres configurados en la Agenda'])
		catalog_sheet.append(['Cargo principal / Otros cargos', 'Usa los cargos configurados en la Agenda'])
		for cell in catalog_sheet[1]:
			cell.font = Font(bold=True)
		buffer = BytesIO()
		workbook.save(buffer)
		buffer.seek(0)
		return send_file(buffer, as_attachment=True, download_name='molde_trabajadores.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	@app.route('/api/horarios/importar-trabajadores', methods=['POST'])
	@login_required
	def horarios_importar_trabajadores():
		if not current_user.can_write('horarios', 'update'):
			return jsonify({'error': 'forbidden'}), 403
		if current_user.rol_nombre != 'admin_general':
			return jsonify({'error': 'Solo Administración General puede importar trabajadores.'}), 403
		archivo = request.files.get('archivo')
		if not archivo:
			return jsonify({'error': 'missing_file'}), 400
		try:
			worksheet = __import__('openpyxl').load_workbook(archivo, read_only=True, data_only=True).active
			rows = list(worksheet.iter_rows(values_only=True))
			if not rows:
				return jsonify({'error': 'empty_file'}), 400
			headers = [str(value or '').strip().lower() for value in rows[0]]
			aliases = {'nombre': 'nombre', 'apellido': 'apellido', 'dni': 'dni', 'telefono': 'telefono', 'fecha nacimiento': 'fechaNacimiento', 'fecha ingreso': 'fechaIngreso', 'direccion': 'direccion', 'emergencia': 'emergenciaNumero', 'grado profesional': 'gradoProfesional', 'profesion': 'profesion', 'institucion de estudios': 'institucionEstudios', 'area': 'area', 'cargo principal': 'cargoPrincipal', 'otros cargos': 'otrosCargos', 'sede': 'sede', 'turno': 'turno', 'dia descanso': 'diaDescanso', 'estado': 'estado'}
			data = []
			for row in rows[1:]:
				values = {aliases[header]: row[index] for index, header in enumerate(headers) if header in aliases and index < len(row)}
				if not str(values.get('nombre') or '').strip() or not str(values.get('apellido') or '').strip():
					continue
				data.append(values)
			registro = db.session.get(AgendaPersistencia, 1)
			payload = json.loads(registro.datos_json) if registro and registro.datos_json else {}
			workers = payload.get('trabajadores', [])
			existing_by_dni = {str(item.get('dni')): item for item in workers if item.get('dni')}
			for item in data:
				dni = str(item.get('dni') or '').strip()
				worker = existing_by_dni.get(dni) if dni else None
				if worker is None:
					worker = {'id': max([int(existing.get('id', 0)) for existing in workers if str(existing.get('id', '')).isdigit()] or [0]) + 1}
					workers.append(worker)
				for key, value in item.items():
					if value is not None and key not in {'sede', 'turno', 'area', 'cargoPrincipal', 'otrosCargos', 'diaDescanso'}:
						worker[key] = str(value)
				worker['estado'] = str(item.get('estado') or worker.get('estado') or 'activo').lower()
				worker['sedeId'] = next((s.id_sede for s in Sede.query.all() if s.nombre_sede.lower().replace('_', ' ') == str(item.get('sede', '')).lower().replace('_', ' ')), worker.get('sedeId'))
				worker['turnoId'] = next((index + 1 for index, turno in enumerate(Turno.query.filter(Turno.id_turno != 'NA').order_by(Turno.nombre_turno).all()) if turno.nombre_turno.lower() == str(item.get('turno', '')).lower()), worker.get('turnoId'))
				areas = payload.get('areas', [])
				cargos = payload.get('cargos', [])
				area_nombre = str(item.get('area') or '').strip().lower()
				cargo_principal = str(item.get('cargoPrincipal') or '').strip().lower()
				otros_nombres = [nombre.strip().lower() for nombre in str(item.get('otrosCargos') or '').split(';') if nombre.strip()]
				worker['areaId'] = next((area.get('id') for area in areas if str(area.get('nombre', '')).lower() == area_nombre), worker.get('areaId'))
				worker['cargos'] = [cargo.get('id') for cargo in cargos if str(cargo.get('nombre', '')).lower() == cargo_principal] or worker.get('cargos', [])
				worker['otrosCargos'] = [cargo.get('id') for cargo in cargos if str(cargo.get('nombre', '')).lower() in otros_nombres]
			payload['trabajadores'] = workers
			if registro is None:
				registro = AgendaPersistencia(id_agenda=1)
				db.session.add(registro)
			registro.datos_json = json.dumps(payload, ensure_ascii=False, default=str)
			registro.actualizado_por = current_user.id_usuario
			db.session.commit()
			return jsonify({'ok': True, 'importados': len(data)})
		except Exception as error:
			return jsonify({'error': str(error)}), 400

	@app.route('/api/horarios/trabajadores')
	@login_required
	def horarios_trabajadores():
		if not current_user.can_view('horarios'):
			return jsonify({'error': 'forbidden'}), 403
		registro = db.session.get(AgendaPersistencia, 1)
		payload = json.loads(registro.datos_json) if registro and registro.datos_json else {}
		workers = payload.get('trabajadores', [])
		if current_user.rol_nombre != 'admin_general':
			workers = [item for item in workers if str(item.get('sedeId')) == str(current_user.id_sede) and str(item.get('turnoId')) in {str(current_user.id_turno), str({'MANANA': 1, 'NOCHE': 2}.get(current_user.id_turno))}]
		return jsonify([{'id': item.get('id'), 'nombre': f"{item.get('nombre', '')} {item.get('apellido', '')}".strip(), 'dni': item.get('dni', '')} for item in workers])

	@app.route('/checklist/template/export')
	@login_required
	def checklist_template_export():
		if not current_user.can_view('checklist'):
			return _forbidden_redirect()

		selected_user_id = request.args.get('f_user', '').strip() or current_user.id_usuario
		if current_user.rol_nombre != 'admin_general' and selected_user_id != current_user.id_usuario:
			selected_user_id = current_user.id_usuario

		selected_user = db.session.get(Usuario, selected_user_id)
		if not selected_user:
			flash('Usuario no encontrado para exportar.', 'error')
			return redirect(url_for('checklist', tab='edit'))

		payload = _build_template_export_payload(selected_user, selected_user_id)
		buffer = BytesIO(json.dumps(payload, ensure_ascii=False, indent=2).encode('utf-8'))
		buffer.seek(0)
		safe_name = (selected_user.username or 'plantilla').replace(' ', '_')
		stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
		return send_file(
			buffer,
			as_attachment=True,
			download_name=f'{safe_name}_plantilla_{stamp}.json',
			mimetype='application/json',
		)

	@app.route('/mermas/productos', methods=['GET'])
	@login_required
	def mermas_productos():
		if not current_user.can_view('mermas'):
			return jsonify({'error': 'forbidden'}), 403
		sede_id = current_user.id_sede
		if current_user.rol_nombre == 'admin_general' and request.args.get('sede', '').isdigit():
			sede_id = int(request.args['sede'])
		rows = db.session.query(Producto, InventarioSede).outerjoin(
			InventarioSede,
			db.and_(InventarioSede.id_producto == Producto.id_producto, InventarioSede.id_sede == sede_id),
		).filter(Producto.estado != 'Inactivo').order_by(Producto.nombre_producto).all()
		return jsonify([{
			'id': producto.id_producto,
			'name': producto.nombre_producto,
			'unit': producto.unidad or 'unidad',
			'cost': float(producto.costo_unitario or 0),
			'stock': float(inventario.stock_actual if inventario else 0),
		} for producto, inventario in rows])

	@app.route('/mermas', methods=['GET', 'POST'])
	@login_required
	def mermas():
		if not current_user.can_view('mermas'):
			return _forbidden_redirect()
		is_admin_general = current_user.rol_nombre == 'admin_general'
		selected_date = _get_selected_app_date()
		month = selected_date.strftime('%Y-%m')
		target_sede = current_user.id_sede
		target_turno = current_user.id_turno
		if is_admin_general and request.args.get('sede', '').isdigit():
			target_sede = int(request.args['sede'])
		if is_admin_general and request.method == 'POST' and request.form.get('sede', '').isdigit():
			target_sede = int(request.form['sede'])
		if is_admin_general and request.args.get('turno'):
			target_turno = request.args['turno']

		if request.method == 'POST':
			if request.form.get('action') == 'catalogo':
				if not is_admin_general:
					return _forbidden_redirect()
				category = request.form.get('categoria', '').strip()
				name = request.form.get('nombre', '').strip()
				if category in {'area', 'turno', 'tipo_merma', 'responsable'} and name:
					if not CatalogoMerma.query.filter_by(categoria=category, nombre=name).first():
						db.session.add(CatalogoMerma(categoria=category, nombre=name))
						db.session.commit()
						flash('Catalogo actualizado.', 'ok')
				return redirect(url_for('mermas'))
			if not current_user.can_write('mermas', 'insert'):
				return _forbidden_redirect()
			try:
				fecha = datetime.strptime(request.form.get('fecha', ''), '%Y-%m-%d').date()
				cantidad = _safe_float(request.form.get('cantidad'), 0)
				producto_id = request.form.get('id_producto', '').strip()
				es_producto_nuevo = producto_id == '__nuevo__'
				if cantidad <= 0:
					raise ValueError('La cantidad debe ser mayor que cero.')
				if es_producto_nuevo:
					nombre_producto = request.form.get('nombre_producto_nuevo', '').strip()
					unidad_nueva = request.form.get('unidad', '').strip() or 'unidad'
					costo = _safe_float(request.form.get('costo_unitario'), 0)
					if not nombre_producto:
						raise ValueError('Escribe el nombre del producto nuevo.')
					producto = Producto(id_producto=_generate_product_id(), nombre_producto=nombre_producto, id_area='', area=_normalize_area(request.form.get('area', '')) or 'cocina', subarea='', unidad=unidad_nueva, costo_unitario=costo, estado='Activo')
					db.session.add(producto)
					db.session.flush()
					inventario = None
					stock = 0.0
				else:
					producto = db.session.get(Producto, producto_id)
					if not producto:
						raise ValueError('Producto invalido.')
					inventario = InventarioSede.query.filter_by(id_sede=target_sede, id_producto=producto.id_producto).with_for_update().first()
					stock = _safe_float(inventario.stock_actual if inventario else 0, 0)
					if not inventario or cantidad > stock:
						raise ValueError(f'Stock insuficiente. Disponible: {stock:.2f}')
					costo = _safe_float(producto.costo_unitario, 0)
				merma = Merma(
					fecha=fecha, mes=fecha.strftime('%Y-%m'), turno=request.form.get('turno', target_turno),
					area=request.form.get('area', '').strip(), id_producto=producto.id_producto,
					tipo_merma=request.form.get('tipo_merma', '').strip(), cantidad=cantidad,
					unidad=producto.unidad or request.form.get('unidad', 'unidad'), costo_unitario=costo,
					costo_total=round(cantidad * costo, 2), responsable=request.form.get('responsable', '').strip(),
					observaciones=request.form.get('observaciones', '').strip(), id_sede=target_sede,
					id_usuario=current_user.id_usuario, bloqueada=True,
				)
				if not merma.area or not merma.tipo_merma or not merma.responsable:
					raise ValueError('Area, tipo de merma y responsable son obligatorios.')
				if inventario:
					inventario.stock_actual = stock - cantidad
				db.session.add(merma)
				if inventario:
					db.session.add(MovimientoInventario(id_sede=target_sede, id_producto=producto.id_producto, cantidad=cantidad, tipo='SALIDA', motivo='Merma', id_usuario=current_user.id_usuario))
				db.session.commit()
				flash('Merma registrada y stock actualizado.', 'ok')
			except (ValueError, IntegrityError) as error:
				db.session.rollback()
				flash(str(error), 'error')
			return redirect(url_for('mermas'))

		catalogs = {key: [row.nombre for row in CatalogoMerma.query.filter_by(categoria=key, activo=True).order_by(CatalogoMerma.nombre).all()] for key in ('area', 'tipo_merma')}
		if not catalogs['area']:
			catalogs['area'] = ['Cocina Fria', 'Emplatado', 'Cocina Caliente', 'Almacen']
		if not catalogs['tipo_merma']:
			catalogs['tipo_merma'] = ['Vencido', 'Malogrado', 'Error preparacion', 'Devolucion', 'Quemado']
		query = Merma.query.filter_by(mes=month)
		if not is_admin_general:
			query = query.filter_by(id_sede=current_user.id_sede, turno=current_user.id_turno)
		else:
			if request.args.get('sede', '').isdigit(): query = query.filter_by(id_sede=int(request.args['sede']))
			if request.args.get('turno'): query = query.filter_by(turno=request.args['turno'])
		for field in ('turno', 'area', 'tipo_merma'):
			value = request.args.get(field, '').strip()
			if value: query = query.filter(getattr(Merma, field) == value)
		from_date_raw = request.args.get('fecha_desde', '').strip()
		to_date_raw = request.args.get('fecha_hasta', '').strip()
		if from_date_raw:
			try: query = query.filter(Merma.fecha >= datetime.strptime(from_date_raw, '%Y-%m-%d').date())
			except ValueError: pass
		if to_date_raw:
			try: query = query.filter(Merma.fecha <= datetime.strptime(to_date_raw, '%Y-%m-%d').date())
			except ValueError: pass
		history = query.order_by(Merma.fecha.desc(), Merma.id_merma.desc()).all()
		catalogs['responsable'] = [row.nombre for row in CatalogoMerma.query.filter_by(categoria='responsable', activo=True).order_by(CatalogoMerma.nombre).all()]
		return render_template('dashboard/mermas.html', history=history, products_url=url_for('mermas_productos', sede=target_sede), catalogs=catalogs, sedes=Sede.query.order_by(Sede.nombre_sede).all(), turnos=Turno.query.order_by(Turno.nombre_turno).all(), target_sede=target_sede, target_turno=target_turno, month=month, is_admin_general=is_admin_general, can_insert=current_user.can_write('mermas', 'insert'), allowed_views=_allowed_views(current_user))

	@app.route('/incidencias', methods=['GET', 'POST'])
	@login_required
	def incidencias():
		if not current_user.can_view('incidencias'):
			return _forbidden_redirect()
		is_admin_general = current_user.rol_nombre == 'admin_general'
		selected_date = _get_selected_app_date()
		month = selected_date.strftime('%Y-%m')
		target_sede = current_user.id_sede
		if is_admin_general and request.args.get('sede', '').isdigit():
			target_sede = int(request.args['sede'])
		if request.method == 'POST':
			if not current_user.can_write('incidencias', 'insert'):
				return _forbidden_redirect()
			try:
				fecha = datetime.strptime(request.form.get('fecha', ''), '%Y-%m-%d').date()
				monto = _safe_float(request.form.get('monto'), 0)
				descuento = request.form.get('descuento') == 'si'
				if fecha.strftime('%Y-%m') != month:
					raise ValueError('Solo puedes registrar datos del mes seleccionado.')
				if not request.form.get('incidencia', '').strip() or not request.form.get('responsable', '').strip() or not request.form.get('encargado', '').strip() or (descuento and monto <= 0):
					raise ValueError('Completa los campos obligatorios y el monto del descuento.')
				if is_admin_general and request.form.get('sede', '').isdigit():
					target_sede = int(request.form['sede'])
				proceso = 'evaluacion' if not is_admin_general else request.form.get('proceso', 'evaluacion').strip().lower()
				if proceso not in {'evaluacion', 'visto', 'aprobado', 'desaprobada'}:
					proceso = 'evaluacion'
				db.session.add(Incidencia(fecha=fecha, mes=fecha.strftime('%Y-%m'), incidencia=request.form['incidencia'].strip(), descripcion=request.form.get('descripcion', '').strip(), responsable=request.form['responsable'].strip(), encargado=request.form['encargado'].strip(), descuento=descuento, monto=monto if descuento else 0, proceso=proceso, id_sede=target_sede, id_usuario=current_user.id_usuario, bloqueada=True))
				db.session.commit()
				flash('Incidencia registrada.', 'ok')
			except (ValueError, IntegrityError) as error:
				db.session.rollback()
				flash(str(error), 'error')
			return redirect(url_for('incidencias'))
		query = Incidencia.query.filter_by(mes=month)
		if not is_admin_general:
			query = query.filter_by(id_sede=current_user.id_sede)
		elif request.args.get('sede', '').isdigit():
			query = query.filter_by(id_sede=target_sede)
		from_date_raw = request.args.get('fecha_desde', '').strip()
		to_date_raw = request.args.get('fecha_hasta', '').strip()
		if from_date_raw:
			try: query = query.filter(Incidencia.fecha >= datetime.strptime(from_date_raw, '%Y-%m-%d').date())
			except ValueError: pass
		if to_date_raw:
			try: query = query.filter(Incidencia.fecha <= datetime.strptime(to_date_raw, '%Y-%m-%d').date())
			except ValueError: pass
		history = query.order_by(Incidencia.fecha.desc(), Incidencia.id_incidencia.desc()).all()
		responsables = Usuario.query.filter_by(id_sede=target_sede).order_by(Usuario.username).all() if not is_admin_general else Usuario.query.order_by(Usuario.username).all()
		return render_template('dashboard/incidencias.html', history=history, month=month, responsables=responsables, sedes=Sede.query.order_by(Sede.nombre_sede).all(), target_sede=target_sede, is_admin_general=is_admin_general, can_insert=current_user.can_write('incidencias', 'insert'), allowed_views=_allowed_views(current_user))

	@app.route('/mermas/<int:id_merma>/editar', methods=['POST'])
	@login_required
	def editar_merma(id_merma):
		if current_user.rol_nombre != 'admin_general':
			return _forbidden_redirect()
		merma = db.session.get(Merma, id_merma)
		if not merma:
			return jsonify({'error': 'not_found'}), 404
		try:
			new_quantity = _safe_float(request.form.get('cantidad'), merma.cantidad)
			if new_quantity <= 0:
				raise ValueError('La cantidad debe ser mayor que cero.')
			inventory = InventarioSede.query.filter_by(id_sede=merma.id_sede, id_producto=merma.id_producto).with_for_update().first()
			available_after_reversal = _safe_float(inventory.stock_actual if inventory else 0) + merma.cantidad
			if not inventory or new_quantity > available_after_reversal:
				raise ValueError('Stock insuficiente para este ajuste.')
			inventory.stock_actual = available_after_reversal - new_quantity
			merma.cantidad = new_quantity
			merma.costo_total = round(new_quantity * merma.costo_unitario, 2)
			for field in ('fecha', 'turno', 'area', 'tipo_merma', 'responsable', 'observaciones'):
				if field in request.form and request.form[field].strip(): setattr(merma, field, request.form[field].strip())
			merma.mes = merma.fecha.strftime('%Y-%m')
			db.session.commit()
			flash('Merma actualizada por Administración General.', 'ok')
		except (ValueError, IntegrityError) as error:
			db.session.rollback()
			flash(str(error), 'error')
		return redirect(url_for('mermas'))

	@app.route('/incidencias/<int:id_incidencia>/editar', methods=['POST'])
	@login_required
	def editar_incidencia(id_incidencia):
		if current_user.rol_nombre != 'admin_general':
			return _forbidden_redirect()
		incidencia = db.session.get(Incidencia, id_incidencia)
		if not incidencia:
			return jsonify({'error': 'not_found'}), 404
		for field in ('fecha', 'incidencia', 'descripcion', 'responsable', 'encargado', 'proceso'):
			if field in request.form and request.form[field].strip():
				value = request.form[field].strip()
				if field == 'fecha': value = datetime.strptime(value, '%Y-%m-%d').date()
				setattr(incidencia, field, value)
		incidencia.descuento = request.form.get('descuento', 'no') == 'si'
		incidencia.monto = _safe_float(request.form.get('monto'), incidencia.monto) if incidencia.descuento else 0
		incidencia.mes = incidencia.fecha.strftime('%Y-%m')
		db.session.commit()
		flash('Incidencia actualizada por Administración General.', 'ok')
		return redirect(url_for('incidencias'))

	@app.route('/arqueo', methods=['GET', 'POST'])
	@login_required
	def arqueo():
		if not current_user.can_view('arqueo'):
			return _forbidden_redirect()
		selected_date = _get_selected_app_date()
		is_admin_general = current_user.rol_nombre == 'admin_general'

		f_sede = request.args.get('sede', '').strip()
		f_turno = request.args.get('turno', '').strip()

		target_sede_id = current_user.id_sede
		target_turno_id = current_user.id_turno
		if is_admin_general:
			if f_sede.isdigit():
				target_sede_id = int(f_sede)
			if f_turno:
				target_turno_id = f_turno

		if not target_sede_id or not target_turno_id:
			flash('No se puede abrir arqueo sin sede y turno definidos.', 'error')
			return redirect(url_for('dashboard'))

		cierre_query = ArqueoCaja.query.filter(
			ArqueoCaja.id_sede == target_sede_id,
			ArqueoCaja.id_turno == target_turno_id,
			ArqueoCaja.fecha == selected_date,
		)
		cierre = cierre_query.order_by(ArqueoCaja.id_arqueo.desc()).first()

		if request.method == 'POST':
			if not current_user.can_write('arqueo', 'update') and not current_user.can_write('arqueo', 'insert'):
				return _forbidden_redirect()

			if cierre is None:
				try:
					cierre = ArqueoCaja(
						id_sede=target_sede_id,
						id_turno=target_turno_id,
						id_usuario=current_user.id_usuario,
						fecha=selected_date,
					)
					db.session.add(cierre)
					db.session.flush()
				except IntegrityError:
					db.session.rollback()
					cierre = ArqueoCaja.query.filter(
						ArqueoCaja.id_sede == target_sede_id,
						ArqueoCaja.id_turno == target_turno_id,
						ArqueoCaja.fecha == selected_date,
					).first()

			if request.is_json:
				payload = request.get_json(silent=True) or {}
			else:
				payload = {'fields': _extract_fields_from_form(request.form)}

			result = _process_arqueo_save(cierre, payload, is_admin_general, current_user, target_sede_id)

			if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
				return jsonify(result)
			flash('Cambios de cierre guardados.', 'ok')
			return redirect(url_for('arqueo', sede=target_sede_id, turno=target_turno_id) if is_admin_general else url_for('arqueo'))

		if cierre and cierre.gastos_json:
			try:
				gastos_actuales = json.loads(cierre.gastos_json)
			except (TypeError, ValueError):
				gastos_actuales = []
		else:
			gastos_actuales = []
		gastos_actuales, gastos_normalizados = _normalizar_gastos_arqueo(gastos_actuales)
		if cierre and gastos_normalizados:
			cierre.gastos_json = json.dumps(gastos_actuales, ensure_ascii=True)
			db.session.commit()

		monto_inicial = _safe_float(cierre.monto_inicial if cierre else 0.0, 0.0)
		pos_tarjetas = _safe_float(cierre.pos_tarjetas if cierre else 0.0, 0.0)
		yape = _safe_float(cierre.yape if cierre else 0.0, 0.0)
		plin = _safe_float(cierre.plin if cierre else 0.0, 0.0)
		efectivo = _safe_float(cierre.efectivo if cierre else 0.0, 0.0)
		venta_sistema = _safe_float(cierre.venta_sistema if cierre else 0.0, 0.0)
		try:
			locked_fields = json.loads(cierre.campos_bloqueados_json or '[]') if cierre else []
		except (TypeError, ValueError):
			locked_fields = []
		if cierre and cierre.efectivo_entregado_guardado:
			locked_fields = sorted(set(locked_fields).union({'efectivo_entregado'}))
		if cierre and cierre.efectivo_dejado_guardado:
			locked_fields = sorted(set(locked_fields).union({'efectivo_dejado_caja_real'}))
		audit_enabled = bool(
			cierre and cierre.venta_sistema_guardada
			and set(locked_fields).intersection({'pos_tarjetas', 'yape', 'plin', 'efectivo'})
		)
		resumen = _calc_cierre_operativo(
			monto_inicial,
			pos_tarjetas,
			yape,
			plin,
			efectivo,
			venta_sistema,
			gastos_actuales,
		)

		historial_query = ArqueoCaja.query.filter(
			ArqueoCaja.fecha == selected_date,
			ArqueoCaja.id_sede == target_sede_id,
			ArqueoCaja.id_turno == target_turno_id,
		)
		if not is_admin_general:
			historial_query = historial_query.filter(
				ArqueoCaja.id_sede == current_user.id_sede,
				ArqueoCaja.id_turno == current_user.id_turno,
			)
		historial_cierres = historial_query.order_by(ArqueoCaja.id_arqueo.desc()).limit(20).all()
		historial_auditoria = []
		if cierre:
			historial_auditoria = _historial_arqueo_por_alcance(cierre)

		sedes_disponibles = []
		turnos_disponibles = []
		target_sede = Sede.query.get(target_sede_id)
		target_turno = Turno.query.get(target_turno_id)
		if is_admin_general:
			sedes_disponibles = Sede.query.order_by(Sede.nombre_sede.asc()).all()
			turnos_disponibles = Turno.query.order_by(Turno.nombre_turno.asc()).all()

		return render_template(
			'dashboard/arqueo_caja.html',
			allowed_views=_allowed_views(current_user),
			cierre=cierre,
			gastos_actuales=gastos_actuales,
			efectivo=efectivo,
			resumen=resumen,
			historial_cierres=historial_cierres,
			historial_auditoria=historial_auditoria,
			target_sede_id=target_sede_id,
			target_turno_id=target_turno_id,
			target_sede_nombre=target_sede.nombre_sede if target_sede else f'Sede {target_sede_id}',
			target_turno_nombre=target_turno.nombre_turno if target_turno else target_turno_id,
			sedes_disponibles=sedes_disponibles,
			turnos_disponibles=turnos_disponibles,
			is_admin_general=is_admin_general,
			locked_fields=locked_fields,
			audit_enabled=audit_enabled,
				can_insert=current_user.can_write('arqueo', 'insert'),
			can_update=current_user.can_write('arqueo', 'update'),
				# Monto inicial base esperado por sede (solo admin_general lo configura)
			expected_base=(target_sede.monto_inicial_base_esperado if target_sede else 0.0),
			loop_month=selected_date.strftime('%Y-%m') if is_admin_general else '',
		)

	@app.route('/arqueo/dashboard', methods=['GET'])
	@login_required
	def arqueo_dashboard():
		if not current_user.can_view('arqueo'):
			return _forbidden_redirect()

		selected_date = _get_selected_app_date()
		is_admin_general = current_user.rol_nombre == 'admin_general'

		rows_query = db.session.query(ArqueoCaja, Sede, Turno).outerjoin(
			Sede, Sede.id_sede == ArqueoCaja.id_sede
		).outerjoin(
			Turno, Turno.id_turno == ArqueoCaja.id_turno
		).filter(
			ArqueoCaja.fecha == selected_date
		)
		if not is_admin_general:
			rows_query = rows_query.filter(
				ArqueoCaja.id_sede == current_user.id_sede,
				ArqueoCaja.id_turno == current_user.id_turno,
			)
		rows = rows_query.order_by(ArqueoCaja.id_arqueo.desc()).all()

		rows_stats = []
		for arqueo, sede, turno in rows:
			monto_inicial = _safe_float(arqueo.monto_inicial, 0.0)
			venta_sistema = _safe_float(arqueo.venta_sistema, 0.0)
			subtotal = _safe_float(arqueo.monto_final, 0.0)
			total_ingresos = (
				_safe_float(arqueo.pos_tarjetas, 0.0)
				+ _safe_float(arqueo.yape, 0.0)
				+ _safe_float(arqueo.plin, 0.0)
				+ _safe_float(arqueo.efectivo, 0.0)
			)
			try:
				gastos_totales = sum(_safe_float(item.get('monto'), 0.0) for item in json.loads(arqueo.gastos_json or '[]'))
			except (TypeError, ValueError):
				gastos_totales = 0.0
			diferencia = (subtotal - monto_inicial) - venta_sistema
			estado = 'Cuadre exacto'
			if diferencia > 0:
				estado = 'Sobrante'
			elif diferencia < 0:
				estado = 'Faltante'
			rows_stats.append({
				'arqueo': arqueo,
				'sede': sede,
				'turno': turno,
				'total_ingresos': total_ingresos,
				'gastos_totales': gastos_totales,
				'subtotal': subtotal,
				'diferencia': diferencia,
				'estado': estado,
			})

		summary = {
			'cierres': len(rows_stats),
			'total_venta_sistema': sum(_safe_float(item['arqueo'].venta_sistema, 0.0) for item in rows_stats),
			'total_ingresos': sum(item['total_ingresos'] for item in rows_stats),
			'total_diferencia': sum(item['diferencia'] for item in rows_stats),
			'sobrantes': sum(1 for item in rows_stats if item['diferencia'] > 0),
			'faltantes': sum(1 for item in rows_stats if item['diferencia'] < 0),
			'cuadrados': sum(1 for item in rows_stats if item['diferencia'] == 0),
		}

		comparacion = None
		if is_admin_general:
			if len(rows_stats) >= 2:
				reference = rows_stats[0]
				target = rows_stats[1]
				comparacion = {
					'reference': reference,
					'target': target,
					'gap_diferencia': reference['diferencia'] - target['diferencia'],
					'gap_venta_sistema': _safe_float(reference['arqueo'].venta_sistema, 0.0) - _safe_float(target['arqueo'].venta_sistema, 0.0),
				}
		else:
			mine = rows_stats[0] if rows_stats else None
			other_row = db.session.query(ArqueoCaja, Sede, Turno).outerjoin(
				Sede, Sede.id_sede == ArqueoCaja.id_sede
			).outerjoin(
				Turno, Turno.id_turno == ArqueoCaja.id_turno
			).filter(
				ArqueoCaja.fecha == selected_date,
				or_(ArqueoCaja.id_sede != current_user.id_sede, ArqueoCaja.id_turno != current_user.id_turno),
			).order_by(ArqueoCaja.id_arqueo.desc()).first()
			if mine and other_row:
				other_arqueo, other_sede, other_turno = other_row
				other_subtotal = _safe_float(other_arqueo.monto_final, 0.0)
				other_diferencia = (other_subtotal - _safe_float(other_arqueo.monto_inicial, 0.0)) - _safe_float(other_arqueo.venta_sistema, 0.0)
				comparacion = {
					'reference': mine,
					'target': {
						'arqueo': other_arqueo,
						'sede': other_sede,
						'turno': other_turno,
						'diferencia': other_diferencia,
					},
					'gap_diferencia': mine['diferencia'] - other_diferencia,
					'gap_venta_sistema': _safe_float(mine['arqueo'].venta_sistema, 0.0) - _safe_float(other_arqueo.venta_sistema, 0.0),
				}

		sede_rollup = {}
		sede_turno_rollup = {}
		for item in rows_stats:
			sede_label = item['sede'].nombre_sede if item['sede'] else f"Sede {item['arqueo'].id_sede}"
			turno_label = item['turno'].nombre_turno if item['turno'] else str(item['arqueo'].id_turno)
			key_sede_turno = f"{sede_label} - {turno_label}"
			if sede_label not in sede_rollup:
				sede_rollup[sede_label] = {
					'venta': 0.0,
					'ingresos': 0.0,
					'pos': 0.0,
					'digital': 0.0,
					'efectivo': 0.0,
					'diferencia': 0.0,
				}
			if key_sede_turno not in sede_turno_rollup:
				sede_turno_rollup[key_sede_turno] = {
					'venta': 0.0,
					'ingresos': 0.0,
					'pos': 0.0,
					'digital': 0.0,
					'efectivo': 0.0,
					'diferencia': 0.0,
				}
			sede_rollup[sede_label]['venta'] += _safe_float(item['arqueo'].venta_sistema, 0.0)
			sede_rollup[sede_label]['ingresos'] += _safe_float(item['total_ingresos'], 0.0)
			sede_rollup[sede_label]['pos'] += _safe_float(item['arqueo'].pos_tarjetas, 0.0)
			sede_rollup[sede_label]['digital'] += _safe_float(item['arqueo'].yape, 0.0) + _safe_float(item['arqueo'].plin, 0.0)
			sede_rollup[sede_label]['efectivo'] += _safe_float(item['arqueo'].efectivo, 0.0)
			sede_rollup[sede_label]['diferencia'] += _safe_float(item['diferencia'], 0.0)

			sede_turno_rollup[key_sede_turno]['venta'] += _safe_float(item['arqueo'].venta_sistema, 0.0)
			sede_turno_rollup[key_sede_turno]['ingresos'] += _safe_float(item['total_ingresos'], 0.0)
			sede_turno_rollup[key_sede_turno]['pos'] += _safe_float(item['arqueo'].pos_tarjetas, 0.0)
			sede_turno_rollup[key_sede_turno]['digital'] += _safe_float(item['arqueo'].yape, 0.0) + _safe_float(item['arqueo'].plin, 0.0)
			sede_turno_rollup[key_sede_turno]['efectivo'] += _safe_float(item['arqueo'].efectivo, 0.0)
			sede_turno_rollup[key_sede_turno]['diferencia'] += _safe_float(item['diferencia'], 0.0)

		bar_labels = sorted(sede_turno_rollup.keys())
		chart_bar = {
			'labels': bar_labels,
			'venta_sistema': [round(sede_turno_rollup[label]['venta'], 2) for label in bar_labels],
			'recaudacion_real': [round(sede_turno_rollup[label]['ingresos'], 2) for label in bar_labels],
		}

		total_pos = sum(item['pos'] for item in sede_rollup.values())
		total_digital = sum(item['digital'] for item in sede_rollup.values())
		total_efectivo = sum(item['efectivo'] for item in sede_rollup.values())
		chart_pie = {
			'labels': ['POS', 'Yape/Plin', 'Efectivo'],
			'values': [round(total_pos, 2), round(total_digital, 2), round(total_efectivo, 2)],
		}

		trend_start = selected_date - timedelta(days=6)
		trend_labels = [
			(trend_start + timedelta(days=i)).strftime('%d/%m')
			for i in range(7)
		]
		trend_lookup = [
			(trend_start + timedelta(days=i)).strftime('%Y-%m-%d')
			for i in range(7)
		]

		trend_query = db.session.query(ArqueoCaja, Sede).outerjoin(
			Sede, Sede.id_sede == ArqueoCaja.id_sede
		).filter(
			ArqueoCaja.fecha >= trend_start,
			ArqueoCaja.fecha <= selected_date,
		)
		if not is_admin_general:
			trend_query = trend_query.filter(
				ArqueoCaja.id_sede == current_user.id_sede,
				ArqueoCaja.id_turno == current_user.id_turno,
			)
		trend_rows = trend_query.all()

		trend_sede_data = defaultdict(lambda: defaultdict(float))
		for arqueo, sede in trend_rows:
			fecha_key = arqueo.fecha.strftime('%Y-%m-%d') if arqueo.fecha else ''
			sede_label = sede.nombre_sede if sede else f"Sede {arqueo.id_sede}"
			total_ing = (
				_safe_float(arqueo.pos_tarjetas, 0.0)
				+ _safe_float(arqueo.yape, 0.0)
				+ _safe_float(arqueo.plin, 0.0)
				+ _safe_float(arqueo.efectivo, 0.0)
			)
			try:
				gastos = sum(_safe_float(item.get('monto'), 0.0) for item in json.loads(arqueo.gastos_json or '[]'))
			except (TypeError, ValueError):
				gastos = 0.0
			trend_sede_data[sede_label][fecha_key] += gastos

		palette = ['#E6C682', '#4A4A4A', '#2D2D2D', '#B98E38', '#7A7A7A']
		chart_trend = {
			'labels': trend_labels,
			'datasets': [],
		}
		for idx, sede_label in enumerate(sorted(trend_sede_data.keys())):
			chart_trend['datasets'].append({
				'label': sede_label,
				'data': [round(trend_sede_data[sede_label].get(day_key, 0.0), 2) for day_key in trend_lookup],
				'borderColor': palette[idx % len(palette)],
				'backgroundColor': palette[idx % len(palette)],
				'tension': 0.25,
				'fill': False,
			})

		benchmark_labels = bar_labels
		benchmark_values = []
		for label in benchmark_labels:
			venta = sede_turno_rollup[label]['venta']
			diferencia = sede_turno_rollup[label]['diferencia']
			operativo = venta + diferencia
			indice = (operativo / venta * 100.0) if venta > 0 else 0.0
			benchmark_values.append(round(indice, 2))
		chart_benchmark = {
			'labels': benchmark_labels,
			'values': benchmark_values,
		}

		comparacion_turnos = []
		sede_items = defaultdict(list)
		for item in rows_stats:
			sede_label = item['sede'].nombre_sede if item['sede'] else f"Sede {item['arqueo'].id_sede}"
			sede_items[sede_label].append(item)

		for sede_label, items in sorted(sede_items.items(), key=lambda entry: entry[0]):
			if len(items) < 2:
				continue
			ordered = sorted(items, key=lambda row: ((row['turno'].nombre_turno if row['turno'] else ''), row['arqueo'].id_turno or ''))
			reference = ordered[0]
			target = ordered[1]
			comparacion_turnos.append({
				'sede': sede_label,
				'reference_turno': reference['turno'].nombre_turno if reference['turno'] else str(reference['arqueo'].id_turno),
				'target_turno': target['turno'].nombre_turno if target['turno'] else str(target['arqueo'].id_turno),
				'reference_diferencia': reference['diferencia'],
				'target_diferencia': target['diferencia'],
				'gap_diferencia': reference['diferencia'] - target['diferencia'],
				'reference_venta': _safe_float(reference['arqueo'].venta_sistema, 0.0),
				'target_venta': _safe_float(target['arqueo'].venta_sistema, 0.0),
				'gap_venta': _safe_float(reference['arqueo'].venta_sistema, 0.0) - _safe_float(target['arqueo'].venta_sistema, 0.0),
			})

		week_start = selected_date - timedelta(days=selected_date.weekday())
		month_start = selected_date.replace(day=1)

		period_query = ArqueoCaja.query.filter(ArqueoCaja.fecha >= month_start, ArqueoCaja.fecha <= selected_date)
		if not is_admin_general:
			period_query = period_query.filter(
				ArqueoCaja.id_sede == current_user.id_sede,
				ArqueoCaja.id_turno == current_user.id_turno,
			)
		period_rows = period_query.all()

		week_ganancia = 0.0
		week_gastos = 0.0
		month_ganancia = 0.0
		month_gastos = 0.0

		for arqueo in period_rows:
			total_ingresos_item = (
				_safe_float(arqueo.pos_tarjetas, 0.0)
				+ _safe_float(arqueo.yape, 0.0)
				+ _safe_float(arqueo.plin, 0.0)
				+ _safe_float(arqueo.efectivo, 0.0)
			)
			try:
				gastos_item = sum(_safe_float(item.get('monto'), 0.0) for item in json.loads(arqueo.gastos_json or '[]'))
			except (TypeError, ValueError):
				gastos_item = 0.0
			ganancia_item = _safe_float(arqueo.monto_final, 0.0) - _safe_float(arqueo.monto_inicial, 0.0)

			if arqueo.fecha and arqueo.fecha >= week_start:
				week_ganancia += ganancia_item
				week_gastos += gastos_item

			month_ganancia += ganancia_item
			month_gastos += gastos_item

		chart_ganancia_gastos = {
			'labels': ['Semana actual', 'Mes actual'],
			'ganancia': [round(week_ganancia, 2), round(month_ganancia, 2)],
			'gastos': [round(week_gastos, 2), round(month_gastos, 2)],
		}

		return render_template(
			'dashboard/arqueo_dashboard.html',
			allowed_views=_allowed_views(current_user),
			selected_date=selected_date,
			is_admin_general=is_admin_general,
			rows_stats=rows_stats,
			summary=summary,
			comparacion=comparacion,
			chart_bar=chart_bar,
			chart_pie=chart_pie,
			chart_trend=chart_trend,
			chart_benchmark=chart_benchmark,
			chart_ganancia_gastos=chart_ganancia_gastos,
			comparacion_turnos=comparacion_turnos,
			sedes_disponibles=Sede.query.order_by(Sede.nombre_sede.asc()).all() if is_admin_general else [],
			turnos_disponibles=Turno.query.order_by(Turno.nombre_turno.asc()).all() if is_admin_general else [],
		)

	@app.route('/arqueo/dashboard/data', methods=['GET'])
	@login_required
	def arqueo_dashboard_data():
		if not current_user.can_view('arqueo'):
			return jsonify({'error': 'forbidden'}), 403
		date_start, date_end = _arqueo_period_params(request.args, current_user)
		sede_ids = [int(value) for value in request.args.getlist('sede') if value.isdigit()] if current_user.rol_nombre == 'admin_general' else None
		turno_ids = request.args.getlist('turno') if current_user.rol_nombre == 'admin_general' else None
		rows = _arqueo_report_rows(date_start, date_end, current_user, sede_ids, turno_ids)
		items = []
		for arqueo, sede, turno, usuario in rows:
			try: gastos = json.loads(arqueo.gastos_json or '[]')
			except (TypeError, ValueError): gastos = []
			total_ingresos, gastos_total, subtotal, operativo, diferencia = _report_metrics(arqueo, gastos)
			items.append({'fecha': arqueo.fecha.isoformat(), 'sede': sede.nombre_sede if sede else str(arqueo.id_sede), 'turno': turno.nombre_turno if turno else arqueo.id_turno, 'ingresos': total_ingresos, 'gastos': gastos_total, 'subtotal': subtotal, 'venta_sistema': _safe_float(arqueo.venta_sistema), 'diferencia': diferencia, 'efectivo_entregado': _safe_float(arqueo.efectivo_entregado), 'efectivo_dejado': _safe_float(arqueo.efectivo_dejado_caja_real)})
		return jsonify({'date_start': date_start.isoformat(), 'date_end': date_end.isoformat(), 'items': items, 'role': current_user.rol_nombre})

	def _arqueo_period_params(args, user):
		today = _get_selected_app_date()
		period = (args.get('period') or 'month').strip()
		start_raw = (args.get('date_start') or '').strip()
		end_raw = (args.get('date_end') or '').strip()
		try:
			date_start = datetime.strptime(start_raw, '%Y-%m-%d').date() if start_raw else None
			date_end = datetime.strptime(end_raw, '%Y-%m-%d').date() if end_raw else None
		except ValueError:
			date_start = date_end = None
		if not date_start or not date_end:
			if period == 'week':
				date_start = today - timedelta(days=today.weekday())
				date_end = date_start + timedelta(days=6)
			elif period == 'last_month':
				first = today.replace(day=1)
				date_end = first - timedelta(days=1)
				date_start = date_end.replace(day=1)
			else:
				date_start = today.replace(day=1)
				date_end = today
		return date_start, date_end


	def _arqueo_report_rows(date_start, date_end, user, sede_ids=None, turno_ids=None):
		query = db.session.query(ArqueoCaja, Sede, Turno, Usuario).outerjoin(
			Sede, Sede.id_sede == ArqueoCaja.id_sede
		).outerjoin(Turno, Turno.id_turno == ArqueoCaja.id_turno).outerjoin(
			Usuario, Usuario.id_usuario == ArqueoCaja.id_usuario
		).filter(ArqueoCaja.fecha >= date_start, ArqueoCaja.fecha <= date_end)
		if user.rol_nombre != 'admin_general':
			query = query.filter(ArqueoCaja.id_sede == user.id_sede, ArqueoCaja.id_turno == user.id_turno)
		else:
			if sede_ids:
				query = query.filter(ArqueoCaja.id_sede.in_(sede_ids))
			if turno_ids:
				query = query.filter(ArqueoCaja.id_turno.in_(turno_ids))
		return query.order_by(ArqueoCaja.fecha.asc(), ArqueoCaja.id_arqueo.asc()).all()


	def _report_metrics(arqueo, gastos):
		total_ingresos = sum(_safe_float(getattr(arqueo, field), 0.0) for field in ('pos_tarjetas', 'yape', 'plin', 'efectivo'))
		gastos_total = sum(_safe_float(item.get('monto'), 0.0) for item in gastos)
		subtotal = total_ingresos + gastos_total
		operativo = subtotal - _safe_float(arqueo.monto_inicial, 0.0)
		diferencia = operativo - _safe_float(arqueo.venta_sistema, 0.0)
		return total_ingresos, gastos_total, subtotal, operativo, diferencia


	@app.route('/arqueo/export_report', methods=['GET'])
	@login_required
	def arqueo_export_report():
		if not current_user.can_view('arqueo'):
			return _forbidden_redirect()
		date_start, date_end = _arqueo_period_params(request.args, current_user)
		sede_ids = [int(value) for value in request.args.getlist('sede') if value.isdigit()] if current_user.rol_nombre == 'admin_general' else None
		turno_ids = request.args.getlist('turno') if current_user.rol_nombre == 'admin_general' else None
		rows = _arqueo_report_rows(date_start, date_end, current_user, sede_ids, turno_ids)
		wb = Workbook()
		ws_summary = wb.active
		ws_summary.title = 'Resumen Ejecutivo'
		ws_detail = wb.create_sheet('Arqueos Detalle')
		ws_expenses = wb.create_sheet('Gastos')
		ws_channels = wb.create_sheet('Ingresos por Canal')
		ws_audit = wb.create_sheet('Auditoria')
		ws_week = wb.create_sheet('Comparativo Semanal')
		ws_month = wb.create_sheet('Comparativo Mensual')
		dark_fill = PatternFill('solid', fgColor='263238')
		money_fmt = '$ #,##0.00'
		for sheet in wb.worksheets:
			sheet.sheet_view.showGridLines = False

		prepared = []
		all_expenses = []
		for arqueo, sede, turno, usuario in rows:
			try:
				gastos = json.loads(arqueo.gastos_json or '[]')
			except (TypeError, ValueError):
				gastos = []
			metrics = _report_metrics(arqueo, gastos)
			prepared.append((arqueo, sede, turno, usuario, gastos, metrics))
			for index, item in enumerate(gastos, 1):
				all_expenses.append((f'{arqueo.id_arqueo}-{index}', arqueo, sede, turno, item, metrics[1]))

		values = [item[5] for item in prepared]
		total_ingresos = sum(item[0] for item in values)
		total_gastos = sum(item[1] for item in values)
		total_subtotal = sum(item[2] for item in values)
		total_operativo = sum(item[3] for item in values)
		total_venta = sum(_safe_float(item[0].venta_sistema, 0.0) for item in prepared)
		total_diferencia = sum(item[4] for item in values)
		total_entregado = sum(_safe_float(item[0].efectivo_entregado, 0.0) for item in prepared)
		total_dejado = sum(_safe_float(item[0].efectivo_dejado_caja_real, 0.0) for item in prepared)
		ws_summary.append([f'Arqueo de Caja - Reporte Periodo {date_start} a {date_end}'])
		ws_summary.append([])
		for label, value in [('Total Ingresos', total_ingresos), ('Total Gastos', total_gastos), ('Subtotal', total_subtotal), ('Total Operativo', total_operativo), ('Venta Sistema total', total_venta), ('Diferencia vs Sistema', total_diferencia), ('Porcentaje de descuadre', total_diferencia / total_venta if total_venta else 0), ('Efectivo Entregado total', total_entregado), ('Efectivo Dejado en Caja total', total_dejado), ('Diferencia promedio efectivo dejado', sum(_safe_float(item[0].diferencia_efectivo_dejado, 0.0) for item in prepared) / len(prepared) if prepared else 0), ('N de cierres', len(prepared))]:
			ws_summary.append([label, value])
		ws_summary.column_dimensions['A'].width = 38
		ws_summary.column_dimensions['B'].width = 22
		ws_summary['A1'].font = Font(size=16, bold=True)
		for row in ws_summary.iter_rows(min_row=3, max_col=2):
			row[0].font = Font(bold=True)
			if isinstance(row[1].value, (int, float)):
				row[1].number_format = '0.00% ' if row[0].value == 'Porcentaje de descuadre' else money_fmt
		ws_summary.conditional_formatting.add('B8', CellIsRule(operator='greaterThan', formula=['0.02'], fill=PatternFill('solid', fgColor='FFC7CE')))

		detail_headers = ['ID Arqueo', 'Fecha', 'Sede', 'Turno', 'Monto Inicial', 'Venta Sistema', 'POS', 'Yape', 'Plin', 'Efectivo', 'Total Ingresos', 'Gastos Totales', 'Subtotal', 'Total Operativo', 'Efectivo Entregado', 'Efectivo Dejado Real', 'Diferencia Efectivo Dejado', 'Diferencia vs Sistema', 'Observaciones']
		ws_detail.append(detail_headers)
		for index, (arqueo, sede, turno, usuario, gastos, metrics) in enumerate(prepared, 2):
			total_ingresos_item, gastos_totales_item, subtotal_item, total_operativo_item, diferencia_item = metrics
			efectivo_dejado_real = _safe_float(getattr(arqueo, 'efectivo_dejado_caja_real', 0.0), 0.0)
			diferencia_efectivo_dejado = _safe_float(getattr(arqueo, 'diferencia_efectivo_dejado', 0.0), 0.0)
			ws_detail.append([arqueo.id_arqueo, arqueo.fecha, sede.nombre_sede if sede else '', turno.nombre_turno if turno else arqueo.id_turno, _safe_float(arqueo.monto_inicial), _safe_float(arqueo.venta_sistema), _safe_float(arqueo.pos_tarjetas), _safe_float(arqueo.yape), _safe_float(arqueo.plin), _safe_float(arqueo.efectivo), total_ingresos_item, gastos_totales_item, subtotal_item, total_operativo_item, _safe_float(arqueo.efectivo_entregado), efectivo_dejado_real, diferencia_efectivo_dejado, diferencia_item, arqueo.observaciones or ''])
		ws_detail.freeze_panes = 'A2'
		ws_detail.auto_filter.ref = f'A1:S{max(ws_detail.max_row, 2)}'
		for column in range(5, 19):
			for cell in ws_detail.iter_cols(min_col=column, max_col=column, min_row=2):
				for value in cell: value.number_format = money_fmt

		ws_expenses.append(['ID Gasto', 'ID Arqueo', 'Fecha', 'Sede', 'Turno', 'Tipo Gasto', 'Nombre', 'Monto', '% del Total Gastos'])
		for expense_id, arqueo, sede, turno, item, total in all_expenses:
			ws_expenses.append([expense_id, arqueo.id_arqueo, arqueo.fecha, sede.nombre_sede if sede else '', turno.nombre_turno if turno else arqueo.id_turno, item.get('tipo', 'Otros'), item.get('nombre', ''), _safe_float(item.get('monto')), _safe_float(item.get('monto')) / total if total else 0])
		ws_expenses.freeze_panes = 'A2'
		ws_channels.append(['ID Arqueo', 'Fecha', 'Sede', 'Turno', 'Canal', 'Monto'])
		for arqueo, sede, turno, usuario, gastos, metrics in prepared:
			for channel in ('pos_tarjetas', 'yape', 'plin', 'efectivo'):
				ws_channels.append([arqueo.id_arqueo, arqueo.fecha, sede.nombre_sede if sede else '', turno.nombre_turno if turno else arqueo.id_turno, channel, _safe_float(getattr(arqueo, channel))])
		ws_audit.append(['ID Historial', 'Fecha Hora', 'Usuario', 'Tipo Evento', 'Campo/Sección', 'Valor Anterior', 'Valor Nuevo'])
		if prepared:
			audit_rows = ArqueoCajaHistorial.query.filter(ArqueoCajaHistorial.id_arqueo.in_([item[0].id_arqueo for item in prepared])).order_by(ArqueoCajaHistorial.fecha_hora.asc()).all()
			for log in audit_rows: ws_audit.append([log.id_historial, log.fecha_hora, log.usuario_id, log.tipo_evento, log.campo_o_seccion_afectada or log.accion, log.valor_anterior, log.valor_nuevo])

		for sheet in (ws_detail, ws_expenses, ws_channels, ws_audit):
			for cell in sheet[1]: cell.font = Font(bold=True, color='FFFFFF'); cell.fill = dark_fill
			for cell in sheet[1]: sheet.column_dimensions[cell.column_letter].width = max(14, min(34, len(str(cell.value)) + 3))
		ws_week.append(['Semana', 'Fecha Inicio', 'Fecha Fin', 'Total Ingresos', 'Total Gastos', 'Subtotal', 'Venta Sistema', 'Diferencia', '% Descuadre', 'Efectivo Entregado', 'Efectivo Dejado'])
		weekly = defaultdict(lambda: [0.0] * 7)
		for arqueo, sede, turno, usuario, gastos, metrics in prepared:
			key = arqueo.fecha.isocalendar().week
			weekly[key][0] += metrics[0]; weekly[key][1] += metrics[1]; weekly[key][2] += metrics[2]; weekly[key][3] += _safe_float(arqueo.venta_sistema); weekly[key][4] += metrics[4]; weekly[key][5] += _safe_float(arqueo.efectivo_entregado); weekly[key][6] += _safe_float(arqueo.efectivo_dejado_caja_real)
		for week, values_week in sorted(weekly.items()):
			ws_week.append([week, '', '', *values_week[:4], values_week[4], values_week[4] / values_week[3] if values_week[3] else 0, values_week[5], values_week[6]])
		ws_month.append(['Mes', 'Total Ingresos', 'Total Gastos', 'Subtotal', 'Venta Sistema', 'Diferencia', '% Descuadre'])
		monthly = defaultdict(lambda: [0.0] * 5)
		for arqueo, sede, turno, usuario, gastos, metrics in prepared:
			key = arqueo.fecha.strftime('%Y-%m'); monthly[key][0] += metrics[0]; monthly[key][1] += metrics[1]; monthly[key][2] += metrics[2]; monthly[key][3] += _safe_float(arqueo.venta_sistema); monthly[key][4] += metrics[4]
		for month, values_month in sorted(monthly.items()): ws_month.append([month, *values_month, values_month[4] / values_month[3] if values_month[3] else 0])
		for sheet in (ws_week, ws_month):
			for cell in sheet[1]: cell.font = Font(bold=True, color='FFFFFF'); cell.fill = dark_fill
			for row in sheet.iter_rows(min_row=2):
				for cell in row[3:]: cell.number_format = '0.00%' if cell.column == 9 or (sheet == ws_month and cell.column == 7) else money_fmt
		# Hoja de Gastos Operativos con detalle del cierre actual
		ws_operativos = wb.create_sheet('Gastos Operativos')
		ws_operativos.append(['Fecha y Hora', 'Categoria / Tipo de Gasto', 'Descripcion / Detalle', 'Monto'])
		operativos_total = 0.0
		for arqueo, sede, turno, usuario in rows:
			fecha_hora = (arq := arqueo.fecha).isoformat() if arqueo.fecha else ''
			for item in _safe_json_list(arqueo.gastos_json):
				monto = _safe_float(item.get('monto'), 0.0)
				operativos_total += monto
				ws_operativos.append([
					fecha_hora,
					item.get('tipo') or 'Otros',
					item.get('nombre') or 'Sin detalle',
					monto,
				])
		ws_operativos.append([])
		ws_operativos.append(['TOTAL GASTOS OPERATIVOS', '', '', operativos_total])
		for cell in ws_operativos[1]:
			cell.font = Font(bold=True, color='FFFFFF')
			cell.fill = dark_fill
		for row in ws_operativos.iter_rows(min_row=2, max_row=ws_operativos.max_row):
			for cell in row[3:4]:
				cell.number_format = money_fmt
		# Usar filtros normales en lugar de tablas estructuradas para mantener
		# compatibilidad con Excel y evitar reparaciones del libro descargado.
		for sheet in (ws_detail, ws_expenses, ws_channels, ws_audit, ws_week, ws_month, ws_operativos):
			if sheet.max_row >= 2 and sheet.max_column >= 1:
				sheet.auto_filter.ref = f'A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}'
		buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
		return send_file(buffer, as_attachment=True, download_name=f'reporte_arqueo_{date_start}_{date_end}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	@app.route('/arqueo/export_month', methods=['GET'])
	@login_required
	def arqueo_export_month():
		# Exporta un Excel con todos los arqueos de un mes, separados por hoja por sede
		if current_user.rol_nombre != 'admin_general':
			return _forbidden_redirect()

		month_param = request.args.get('month', '').strip()
		if not month_param:
			# por defecto mes actual
			now = datetime.utcnow()
			month_param = now.strftime('%Y-%m')

		try:
			year_str, mon_str = month_param.split('-')
			year = int(year_str)
			month = int(mon_str)
		except Exception:
			flash('Formato de mes inválido. Use YYYY-MM.', 'error')
			return redirect(url_for('arqueo'))

		from calendar import monthrange
		first_day = datetime(year, month, 1).date()
		last_day = datetime(year, month, monthrange(year, month)[1]).date()

		rows = db.session.query(ArqueoCaja, Sede, Turno, Usuario).outerjoin(
			Sede, Sede.id_sede == ArqueoCaja.id_sede
		).outerjoin(
			Turno, Turno.id_turno == ArqueoCaja.id_turno
		).outerjoin(
			Usuario, Usuario.id_usuario == ArqueoCaja.id_usuario
		).filter(
			ArqueoCaja.fecha >= first_day,
			ArqueoCaja.fecha <= last_day,
		).order_by(Sede.nombre_sede.asc(), ArqueoCaja.fecha.asc()).all()

		# Agrupar por sede
		by_sede = defaultdict(list)
		for arqueo, sede, turno, usuario in rows:
			by_sede[sede.nombre_sede if sede else 'Sin Sede'].append((arqueo, sede, turno, usuario))

		wb = Workbook()
		# eliminar sheet por defecto
		default = wb.active
		wb.remove(default)

		for sede_name, items in by_sede.items():
			ws = wb.create_sheet(title=(sede_name[:31] or 'Sede'))
			headers = [
				'ID Arqueo', 'Sede', 'Turno', 'Fecha', 'Usuario (cierre)', 'Monto inicial', 'Venta sistema', 'Monto final',
				'POS tarjetas', 'Yape', 'Plin', 'Efectivo', 'Efectivo Dejado Real', 'Diferencia Efectivo Dejado', 'Gastos (JSON)', 'Observaciones'
			]
			ws.append(headers)
			for arqueo, sede, turno, usuario in items:
				ws.append([
					arqueo.id_arqueo,
					sede.nombre_sede if sede else '',
					(turno.nombre_turno if turno else arqueo.id_turno),
					(arqueo.fecha.isoformat() if arqueo.fecha else ''),
					(usuario.username if usuario else arqueo.id_usuario),
					_safe_float(arqueo.monto_inicial, 0.0),
					_safe_float(arqueo.venta_sistema, 0.0),
					_safe_float(arqueo.monto_final, 0.0),
					_safe_float(arqueo.pos_tarjetas, 0.0),
					_safe_float(arqueo.yape, 0.0),
					_safe_float(arqueo.plin, 0.0),
					_safe_float(arqueo.efectivo, 0.0),
					_safe_float(getattr(arqueo, 'efectivo_dejado_caja_real', 0.0), 0.0),
					_safe_float(getattr(arqueo, 'diferencia_efectivo_dejado', 0.0), 0.0),
					(arqueo.gastos_json or ''),
					(arqueo.observaciones or ''),
				])

		buffer = BytesIO()
		wb.save(buffer)
		buffer.seek(0)
		stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
		filename = f'arqueos_{year:04d}_{month:02d}_{stamp}.xlsx'
		return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	@app.route('/admin/ajustes', methods=['GET', 'POST'])
	@login_required
	def ajustes():
		if not current_user.can_view('ajustes'):
			return _forbidden_redirect()

		if request.method == 'POST':
			if not current_user.can_write('ajustes', 'insert'):
				return _forbidden_redirect()

			tipo_form = request.form.get('tipo_form')
			if tipo_form == 'sede':
				nombre_sede = request.form.get('nombre_sede', '').strip()
				if not nombre_sede:
					flash('Nombre de sede requerido.', 'error')
					return redirect(url_for('ajustes'))
				db.session.add(Sede(nombre_sede=nombre_sede))
			elif tipo_form == 'update_sede':
				try:
					sede_id = int(request.form.get('id_sede', ''))
				except (TypeError, ValueError):
					flash('Sede invalida.', 'error')
					return redirect(url_for('ajustes'))
				nombre_sede = request.form.get('nombre_sede', '').strip()
				sede = Sede.query.get(sede_id)
				if not sede or not nombre_sede:
					flash('La sede y su nombre son obligatorios.', 'error')
					return redirect(url_for('ajustes'))
				sede.nombre_sede = nombre_sede
			elif tipo_form == 'delete_sede':
				try:
					sede_id = int(request.form.get('id_sede', ''))
				except (TypeError, ValueError):
					flash('Sede invalida.', 'error')
					return redirect(url_for('ajustes'))
				sede = Sede.query.get(sede_id)
				if not sede:
					flash('Sede no encontrada.', 'error')
					return redirect(url_for('ajustes'))
				in_use = any((
					Usuario.query.filter_by(id_sede=sede_id).first(),
					InventarioSede.query.filter_by(id_sede=sede_id).first(),
					ChecklistPedido.query.filter_by(id_sede=sede_id).first(),
					ArqueoCaja.query.filter_by(id_sede=sede_id).first(),
					RecordatorioCierre.query.filter_by(id_sede=sede_id).first(),
				))
				if in_use:
					flash('No se puede eliminar una sede con datos asociados. Puedes editar su nombre.', 'error')
					return redirect(url_for('ajustes'))
				db.session.delete(sede)
			elif tipo_form == 'recordatorio_cierre':
				try:
					id_sede = int(request.form.get('id_sede'))
				except (TypeError, ValueError):
					flash('Sede invalida.', 'error')
					return redirect(url_for('ajustes'))
				id_turno = request.form.get('id_turno', '').strip()
				hora_cierre = request.form.get('hora_cierre', '').strip()
				if not id_turno or not re.match(r'^([01]\d|2[0-3]):[0-5]\d$', hora_cierre):
					flash('Hora de cierre invalida.', 'error')
					return redirect(url_for('ajustes'))
				recordatorio = RecordatorioCierre.query.filter_by(id_sede=id_sede, id_turno=id_turno).first()
				if not recordatorio:
					recordatorio = RecordatorioCierre(id_sede=id_sede, id_turno=id_turno)
					db.session.add(recordatorio)
				recordatorio.hora_cierre = hora_cierre
				recordatorio.activo = request.form.get('activo') == 'on'
			elif tipo_form == 'update_recordatorio':
				try:
					recordatorio_id = int(request.form.get('id_recordatorio', ''))
				except (TypeError, ValueError):
					flash('Recordatorio invalido.', 'error')
					return redirect(url_for('ajustes'))
				recordatorio = RecordatorioCierre.query.get(recordatorio_id)
				hora_cierre = request.form.get('hora_cierre', '').strip()
				id_turno = request.form.get('id_turno', '').strip()
				if not recordatorio or not id_turno or not re.match(r'^([01]\d|2[0-3]):[0-5]\d$', hora_cierre):
					flash('Datos invalidos para el recordatorio.', 'error')
					return redirect(url_for('ajustes'))
				recordatorio.id_turno = id_turno
				recordatorio.hora_cierre = hora_cierre
				recordatorio.activo = request.form.get('activo') == 'on'
			elif tipo_form == 'delete_recordatorio':
				try:
					recordatorio_id = int(request.form.get('id_recordatorio', ''))
				except (TypeError, ValueError):
					flash('Recordatorio invalido.', 'error')
					return redirect(url_for('ajustes'))
				recordatorio = RecordatorioCierre.query.get(recordatorio_id)
				if recordatorio:
					db.session.delete(recordatorio)
			elif tipo_form == 'usuario':
				new_id = request.form.get('id_usuario', '').strip()
				new_username = request.form.get('username', '').strip()
				if not new_id or not new_username:
					flash('ID y username son obligatorios.', 'error')
					return redirect(url_for('ajustes'))
				if Usuario.query.filter_by(id_usuario=new_id).first():
					flash('El ID de usuario ya existe.', 'error')
					return redirect(url_for('ajustes'))
				if Usuario.query.filter(db.func.lower(Usuario.username) == new_username.lower()).first():
					flash('El username ya existe.', 'error')
					return redirect(url_for('ajustes'))
				db.session.add(
					Usuario(
						id_usuario=new_id,
						username=new_username,
						password_hash=generate_password_hash(request.form.get('password', '123456')),
						id_rol=int(request.form.get('id_rol')),
						id_sede=int(request.form.get('id_sede')),
						id_turno=request.form.get('id_turno'),
					)
				)
			elif tipo_form == 'update_usuario':
				old_id = request.form.get('old_id_usuario', '').strip()
				new_id = request.form.get('id_usuario', '').strip()
				new_username = request.form.get('username', '').strip()
				new_password = request.form.get('password', '')

				if not old_id or not new_id or not new_username:
					flash('ID actual, nuevo ID y username son obligatorios.', 'error')
					return redirect(url_for('ajustes'))

				usuario = Usuario.query.filter_by(id_usuario=old_id).first()
				if not usuario:
					flash('No se encontro el usuario a actualizar.', 'error')
					return redirect(url_for('ajustes'))

				id_in_use = Usuario.query.filter(Usuario.id_usuario == new_id, Usuario.id_usuario != old_id).first()
				if id_in_use:
					flash('El nuevo ID ya esta en uso.', 'error')
					return redirect(url_for('ajustes'))

				username_in_use = Usuario.query.filter(
					db.func.lower(Usuario.username) == new_username.lower(),
					Usuario.id_usuario != old_id,
				).first()
				if username_in_use:
					flash('El username ya esta en uso.', 'error')
					return redirect(url_for('ajustes'))

				if new_id != old_id:
					replacement = Usuario(
						id_usuario=new_id,
						username=new_username,
						password_hash=generate_password_hash(new_password) if new_password else usuario.password_hash,
						id_rol=int(request.form.get('id_rol', usuario.id_rol)),
						id_sede=int(request.form.get('id_sede', usuario.id_sede)),
						id_turno=request.form.get('id_turno', usuario.id_turno),
					)
					db.session.add(replacement)
					db.session.flush()

					ChecklistPedido.query.filter_by(id_usuario=old_id).update({'id_usuario': new_id})
					MovimientoInventario.query.filter_by(id_usuario=old_id).update({'id_usuario': new_id})
					ArqueoCaja.query.filter_by(id_usuario=old_id).update({'id_usuario': new_id})
					PlantillaChecklistItem.query.filter_by(id_usuario=old_id).update({'id_usuario': new_id})
					DetallePedido.query.filter_by(id_usuario=old_id).update({'id_usuario': new_id})

					db.session.delete(usuario)
				else:
					usuario.username = new_username
					usuario.id_rol = int(request.form.get('id_rol', usuario.id_rol))
					usuario.id_sede = int(request.form.get('id_sede', usuario.id_sede))
					usuario.id_turno = request.form.get('id_turno', usuario.id_turno)
					if new_password:
						usuario.password_hash = generate_password_hash(new_password)

				flash('Usuario actualizado correctamente.', 'ok')
				db.session.commit()
				return redirect(url_for('ajustes'))
			elif tipo_form == 'delete_usuario':
				user_id = request.form.get('id_usuario', '').strip()
				if not user_id:
					flash('Usuario invalido.', 'error')
					return redirect(url_for('ajustes'))
				if user_id == current_user.id_usuario:
					flash('No puedes eliminar tu propio usuario mientras estas logueado.', 'error')
					return redirect(url_for('ajustes'))

				usuario = Usuario.query.filter_by(id_usuario=user_id).first()
				if not usuario:
					flash('No se encontro el usuario.', 'error')
					return redirect(url_for('ajustes'))

				ChecklistPedido.query.filter_by(id_usuario=user_id).update({'id_usuario': None})
				MovimientoInventario.query.filter_by(id_usuario=user_id).update({'id_usuario': None})
				ArqueoCaja.query.filter_by(id_usuario=user_id).update({'id_usuario': None})
				PlantillaChecklistItem.query.filter_by(id_usuario=user_id).delete()
				DetallePedido.query.filter_by(id_usuario=user_id).update({'id_usuario': None})
				db.session.delete(usuario)
				db.session.commit()
				flash('Usuario eliminado correctamente.', 'ok')
				return redirect(url_for('ajustes'))
			elif tipo_form == 'sede_base':
				# Actualizar monto inicial base esperado para una sede (solo admin_general lo hace)
				if current_user.rol_nombre != 'admin_general':
					return _forbidden_redirect()
				sede_id = request.form.get('id_sede')
				valor = request.form.get('monto_inicial_base_esperado', '').strip()
				try:
					valor_f = float(valor) if valor != '' else 0.0
				except ValueError:
					flash('Valor invalido para monto inicial base.', 'error')
					return redirect(url_for('ajustes'))
				try:
					sede_id_int = int(sede_id)
				except Exception:
					sede_id_int = None
				sede = Sede.query.filter_by(id_sede=sede_id_int).first()
				if not sede:
					flash('Sede no encontrada.', 'error')
					return redirect(url_for('ajustes'))
				sede.monto_inicial_base_esperado = valor_f

			db.session.commit()
			flash('Configuracion guardada.', 'ok')

		sedes = Sede.query.order_by(Sede.nombre_sede).all()
		roles = Rol.query.order_by(Rol.nombre_rol).all()
		turnos = Turno.query.order_by(Turno.nombre_turno).all()
		usuarios = Usuario.query.order_by(Usuario.username).all()
		recordatorios = RecordatorioCierre.query.order_by(RecordatorioCierre.id_sede, RecordatorioCierre.id_turno).all()
		return render_template(
			'admin/ajustes.html',
			allowed_views=_allowed_views(current_user),
			sedes=sedes,
			roles=roles,
			turnos=turnos,
			usuarios=usuarios,
			recordatorios=recordatorios,
			sede_names={s.id_sede: s.nombre_sede for s in sedes},
			turno_names={t.id_turno: t.nombre_turno for t in turnos},
		)

	@app.context_processor
	def inject_globals():
		selected_date = session.get('app_date', '').strip()
		if not selected_date:
			app_date = _get_operation_date()
			selected_date = app_date.strftime('%Y-%m-%d')
			session['app_date'] = selected_date
		else:
			try:
				app_date = datetime.strptime(selected_date, '%Y-%m-%d')
			except ValueError:
				app_date = _get_operation_date()
				selected_date = app_date.strftime('%Y-%m-%d')
				session['app_date'] = selected_date
		reminder = None
		if current_user.is_authenticated and current_user.rol_nombre == 'admin_sala' and current_user.id_sede and current_user.id_turno:
			reminder = RecordatorioCierre.query.filter_by(id_sede=current_user.id_sede, id_turno=current_user.id_turno, activo=True).first()
		return {
			'allowed_views': _allowed_views(current_user) if current_user.is_authenticated else [],
			'today_text': app_date.strftime('%d/%m/%Y'),
			'today_value': selected_date,
			'current_date_obj': app_date.date(),
			'cierre_reminder': {'time': reminder.hora_cierre, 'sede': reminder.id_sede, 'turno': reminder.id_turno} if reminder else None,
		}

	return app


app = create_app()


if __name__ == '__main__':
	app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG', '0') == '1')
